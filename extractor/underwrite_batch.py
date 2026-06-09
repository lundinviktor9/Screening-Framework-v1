#!/usr/bin/env python3
"""
underwrite_batch.py - run the rent-roll normaliser over many broker files at once, for a
live-LLM hardening / review pass. For each file it prints which path resolved the layout
(hand adapter / known layout / LLM), the proposed column map, 3 sample PARSED rows, and the
Mode A judgement flags - so you can eyeball mapping quality before trusting the auto path.

This is the non-UI way to do the `propose_mapping` runs. It needs ANTHROPIC_API_KEY in the
environment (or extractor/.env) ONLY for genuinely unknown layouts; hand/known layouts run
offline. It does NOT run Mode B (no LibreOffice needed) - it stops at the canonical RR.

Usage:
    python -m extractor.underwrite_batch <folder-or-files...> [--asset NAME] [--region NAME]
                                         [--out DIR] [--sheet SHEETNAME]

Examples:
    python -m extractor.underwrite_batch deals_inbox/rent_rolls
    python -m extractor.underwrite_batch a.xlsx b.xlsx --region "Milton Keynes"
"""
from __future__ import annotations
import argparse
import json
import sys
from pathlib import Path

from dotenv import load_dotenv

REPO_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(REPO_ROOT))
load_dotenv(Path(__file__).parent / ".env")

from underwrite.engine.normalisers import normalise_auto as na   # noqa: E402
from underwrite import adapter as uw                              # noqa: E402


def _gather(paths):
    files = []
    for p in paths:
        pp = Path(p)
        if pp.is_dir():
            files += sorted(pp.glob("*.xlsx")) + sorted(pp.glob("*.xls")) + sorted(pp.glob("*.csv"))
        elif pp.exists():
            files.append(pp)
        else:
            print(f"  ! not found: {p}")
    return files


def run(paths, asset=None, region=None, out_dir="underwrite_runs/batch", sheet=None):
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    files = _gather(paths)
    if not files:
        print("No input files.")
        return
    client = None
    summary = []
    for f in files:
        print("\n" + "=" * 78)
        print(f"FILE: {f.name}")
        rr = str(out / f"{f.stem}__canonical.xlsx")
        try:
            if client is None:
                try:
                    import anthropic
                    client = anthropic.Anthropic()
                except Exception:
                    client = None  # offline; hand/known layouts still work
            res = na.normalise_auto(str(f), asset or f.stem, region or "UK", rr,
                                    rr_sheet="DealRR", anthropic_client=client, sheet=sheet)
        except Exception as e:
            print(f"  ! normalise failed: {e}")
            summary.append((f.name, "ERROR", 0, str(e)))
            continue
        mp = res["mapping"]
        src = mp.get("_source", "llm")
        print(f"  resolved by : {src}")
        if mp.get("header_row"):
            print(f"  header row  : {mp.get('header_row')}  data from {mp.get('data_start_row')}")
        cols = mp.get("columns") or {}
        if cols:
            print("  column map  :")
            for k, v in cols.items():
                print(f"      {k:28} <- col {v}")
        # 3 sample parsed rows from the canonical RR
        import openpyxl
        from openpyxl.utils import column_index_from_string as cix
        ws = openpyxl.load_workbook(rr, data_only=True)["DealRR"]
        hdr = {c: (ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)}
        print("  sample rows :")
        for r in range(2, 5):
            if ws.cell(r, cix("E")).value in (None, ""):
                break
            row = {hdr[c]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)
                   if hdr[c] and ws.cell(r, c).value not in (None, "")}
            print("      " + json.dumps(row, default=str)[:200])
        # Mode A flags
        try:
            ma = uw.run_mode_a(f.stem, rr, "DealRR")
            print(f"  units       : {ma['units']} | schema_errors: {ma['schema_errors']}")
            for fl in (res.get("flags") or []) + (ma.get("flags") or []):
                print(f"      FLAG {fl.get('unit','') } {fl.get('signal') or fl.get('field')}: {fl.get('note','')}")
            summary.append((f.name, src, ma["units"], "ok"))
        except Exception as e:
            print(f"  ! Mode A failed: {e}")
            summary.append((f.name, src, res["units"], f"modeA:{e}"))
        print(f"  canonical RR -> {rr}")

    print("\n" + "=" * 78)
    print("SUMMARY")
    for name, src, units, note in summary:
        print(f"  {name:40} {src:16} units={units:<4} {note}")
    print(f"\nReview the column maps + sample rows above. Promote any broker you'll see again to a")
    print(f"hand adapter in underwrite/engine/normalisers/hand_adapters.py (see Cannon/Meadow).")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("paths", nargs="+", help="folder(s) or file(s)")
    ap.add_argument("--asset", default=None)
    ap.add_argument("--region", default=None)
    ap.add_argument("--out", default="underwrite_runs/batch")
    ap.add_argument("--sheet", default=None)
    a = ap.parse_args()
    run(a.paths, a.asset, a.region, a.out, a.sheet)
