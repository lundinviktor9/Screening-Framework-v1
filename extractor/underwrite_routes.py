"""
extractor/underwrite_routes.py - FastAPI routes for the MLI underwrite stage (Mode A/B).

Wires underwrite/adapter.py (Mode A/B) + underwrite/engine/normalisers/normalise_auto.py
into the existing pipeline server and stamps results onto the deal record via
DealStore.update(deal_id, {"underwrite": {...}}).

Flow (see underwrite/INTERFACE.md):
  upload rent roll
    -> normalise_auto: hand-adapter | known-map | LLM-map (layout) + apply (real cells)
    -> run_mode_a: validate + surface judgement flags (each with a stored RATIONALE)  [flagged]
  (analyst reviews mapping + 3 sample rows + flags, optionally corrects the column map)
    -> POST /confirm-mapping: re-run apply_mapping idempotently
  (analyst signs off mapping AND flags, sets assumptions, optionally a change note)
    -> POST /run: run_mode_b -> returns + checks; appended as a VERSIONED run         [underwritten]

Audit trail (this module's job, beyond the engine):
  - Every judgement flag carries a `rationale` = WHY it is a judgement call (not just that it
    fired), plus the analyst's `resolution` once signed off.
  - The FIRST successful run is frozen as `baseline` = "what the machine produced".
  - Every run is appended to `runs[]` with: the exact `assumptions` used, a `vs_baseline` and
    `vs_previous` assumption diff, the resulting returns/checks, and an analyst `note` recording
    what was changed and why. Nothing is overwritten - you can always see how a number moved.

Rules honoured (INTERFACE.md s.5): deal_id threads through; human-in-the-loop mandatory (/run
refuses without mapping + flag sign-off); no false precision (display_returns only when
checks.pass); values come from cells not the LLM; outputs OUTSIDE OneDrive (UNDERWRITE_OUT_DIR).

Exposes a factory `make_underwrite_router(store)` so it shares the single DealStore instance.
"""
from __future__ import annotations

import os
import sys
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.utils import column_index_from_string as cix

REPO_ROOT = Path(__file__).parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from underwrite import adapter as uw                                  # noqa: E402
from underwrite.engine.normalisers import normalise_auto as na        # noqa: E402

# ---------------------------------------------------------------- config

RR_SHEET = "DealRR"
MAP_MODEL = os.environ.get("UNDERWRITE_MAP_MODEL", "claude-sonnet-4-6")

# Outputs MUST live outside OneDrive (sync dehydrates/corrupts xlsx mid-write).
UNDERWRITE_DIR = Path(
    os.environ.get("UNDERWRITE_OUT_DIR", str(REPO_ROOT / "extractor" / "underwrite_runs"))
)
UNDERWRITE_DIR.mkdir(parents=True, exist_ok=True)

# Why each judgement signal/field is a JUDGEMENT CALL (recorded next to every flag so the
# rationale survives, not just the fact that it fired). Keyed by signal, then by field.
FLAG_RATIONALE = {
    "vacant@entry": "Vacant units carry no contractual income; value depends on the analyst's "
                    "view of headline/guarantee rent and the re-letting (void/rent-free/capex) "
                    "assumptions - a pricing judgement, not a fact in the schedule.",
    "under_offer": "An agreed-but-not-completed letting; whether and when to recognise the income "
                   "is a judgement on deal certainty.",
    "break_taken": "Whether the tenant exercises the break is an assumption about future behaviour "
                   "that materially changes the cashflow and re-letting profile.",
    "break_date": "Whether the tenant exercises the break is an assumption about future behaviour "
                  "that materially changes the cashflow and re-letting profile.",
    "vacate@expiry": "Assuming the tenant vacates at expiry (vs renews) drives void/re-let costs "
                     "and the income profile - a forward-looking judgement.",
}
FIELD_RATIONALE = {
    "Entry Yield (NIY)": "The acquisition yield is the PRICING decision; the schedule NIY is a "
                         "starting point, not the agreed price. Confirm, don't inherit.",
    "Exit Yield": "The exit yield is the single biggest value driver and is a forward-looking "
                  "market view set by the analyst.",
    "Vacant @ Entry": "Confirm the vacant valuation basis (capitalise guarantee vs headline) - it "
                      "changes the entry price.",
}

# Gap-overrideable schedule fields: re-letting / valuation inputs that brokers often leave
# blank per-unit and which legitimately take a deal-level default. Matched by DealRR COLUMN
# LETTER (the canonical schema is fixed; header text carries a non-ASCII '£' that we avoid).
# kind: "pct" values are entered as percentages in the UI and stored as decimals (÷100).
GAP_FIELDS = [
    {"field": "Exit Yield", "column": "I", "kind": "pct", "unit": "%"},
    {"field": "ERV (psf)", "column": "U", "kind": "num", "unit": "£ psf"},
    {"field": "Assumed Void (mths)", "column": "Y", "kind": "num", "unit": "mths"},
    {"field": "Assumed Rent Free (mths)", "column": "Z", "kind": "num", "unit": "mths"},
    {"field": "Re-letting Capex (psf)", "column": "AA", "kind": "num", "unit": "£ psf"},
    {"field": "Term Certain (mths)", "column": "BC", "kind": "num", "unit": "mths"},
]
_GAP_BY_FIELD = {g["field"]: g for g in GAP_FIELDS}


# ---------------------------------------------------------------- request bodies

class ConfirmMappingBody(BaseModel):
    mapping: Dict[str, Any]
    asset: Optional[str] = None
    region: Optional[str] = None
    note: Optional[str] = None
    analyst: Optional[str] = None


class RunBody(BaseModel):
    assumptions: Dict[str, Any] = {}
    gap_overrides: Dict[str, Any] = {}   # field -> deal-level default for blank schedule cells
    flag_resolutions: Optional[List[Dict[str, Any]]] = None
    mapping_signed_off: bool = False
    flags_signed_off: bool = False
    asset: Optional[str] = None
    region: Optional[str] = None
    note: Optional[str] = None       # what changed vs the machine output, and why
    analyst: Optional[str] = None    # who ran it


# ---------------------------------------------------------------- helpers

def _now() -> str:
    return datetime.utcnow().isoformat() + "Z"


def _deal_or_404(store, deal_id: str) -> Dict[str, Any]:
    deal = store.read_by_id(deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal not found: {deal_id}")
    return deal


def _defaults(deal: Dict[str, Any]) -> Dict[str, str]:
    ef = deal.get("extracted_fields") or {}
    asset = ef.get("Project Name") or ef.get("Address") or deal.get("source_filename") or deal["deal_id"]
    region = ef.get("Location") or (deal.get("market_ids") or [""])[0] or "UK"
    return {"asset": str(asset), "region": str(region)}


def _merge_underwrite(store, deal_id: str, patch: Dict[str, Any]) -> Dict[str, Any]:
    """Read-merge-write the `underwrite` sub-object (DealStore.update is a shallow dict.update)."""
    deal = _deal_or_404(store, deal_id)
    block = dict(deal.get("underwrite") or {})
    block.update(patch)
    block["updated_at"] = _now()
    store.update(deal_id, {"underwrite": block})
    return block


def _deal_dir(deal_id: str) -> Path:
    d = UNDERWRITE_DIR / deal_id
    d.mkdir(parents=True, exist_ok=True)
    return d


def _enrich_flags(flags: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Attach a stable `rationale` (the WHY) to each flag, keyed by signal then field."""
    out = []
    for f in flags or []:
        g = dict(f)
        if "rationale" not in g:
            sig = g.get("signal")
            fld = g.get("field")
            g["rationale"] = FLAG_RATIONALE.get(sig) or FIELD_RATIONALE.get(fld) or (
                "Analyst judgement required; confirm before underwriting.")
        out.append(g)
    return out


def _flag_key(f: Dict[str, Any]) -> str:
    return f"{f.get('unit', '')}|{f.get('signal') or f.get('field') or ''}"


def _merge_flag_resolutions(flags: List[Dict[str, Any]],
                            resolutions: Optional[List[Dict[str, Any]]],
                            analyst: Optional[str]) -> List[Dict[str, Any]]:
    """Stamp analyst resolutions onto matching flags (by unit+signal/field)."""
    if not resolutions:
        return flags
    by_key = {}
    for r in resolutions:
        by_key[_flag_key(r)] = r
    out = []
    for f in flags:
        g = dict(f)
        r = by_key.get(_flag_key(f))
        if r:
            g["resolution"] = {
                "decision": r.get("decision"),
                "note": r.get("note"),
                "analyst": analyst,
                "at": _now(),
            }
        out.append(g)
    return out


def _assumptions_diff(old: Optional[Dict[str, Any]], new: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    """{key: {"from": old, "to": new}} for keys that changed (added/removed/edited)."""
    old = old or {}
    new = new or {}
    diff = {}
    for k in set(old) | set(new):
        if old.get(k) != new.get(k):
            diff[k] = {"from": old.get(k), "to": new.get(k)}
    return diff


def _sample_rows(rr_xlsx: str, rr_sheet: str = RR_SHEET, n: int = 3) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(rr_xlsx, data_only=True)
    if rr_sheet not in wb.sheetnames:
        return []
    ws = wb[rr_sheet]
    headers = {c: (ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)}
    out: List[Dict[str, Any]] = []
    for r in range(2, min(ws.max_row, 1 + n) + 1):
        if ws.cell(r, cix("E")).value in (None, ""):
            break
        row = {headers[c]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)
               if headers[c] and ws.cell(r, c).value not in (None, "")}
        out.append(row)
    return out


def _rr_data_rows(ws) -> List[int]:
    """Row numbers carrying a unit (column E / Unit Number non-empty)."""
    e = cix("E")
    rows = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, e).value not in (None, ""):
            rows.append(r)
    return rows


def _compute_gaps(rr_xlsx: Optional[str], rr_sheet: str = RR_SHEET) -> List[Dict[str, Any]]:
    """Per-field count of blank cells across the canonical RR's unit rows.

    Returns one entry per GAP_FIELD: {field, column, kind, unit, missing, total}. `missing`==0
    means the schedule is complete for that field (the UI disables its override input)."""
    if not rr_xlsx or not Path(rr_xlsx).exists():
        return []
    try:
        wb = openpyxl.load_workbook(rr_xlsx, data_only=True)
    except Exception:
        return []
    if rr_sheet not in wb.sheetnames:
        return []
    ws = wb[rr_sheet]
    data_rows = _rr_data_rows(ws)
    total = len(data_rows)
    out: List[Dict[str, Any]] = []
    for g in GAP_FIELDS:
        col = cix(g["column"])
        missing = sum(1 for r in data_rows if ws.cell(r, col).value in (None, ""))
        out.append({**g, "missing": missing, "total": total})
    return out


def _apply_gap_overrides(rr_xlsx: str, overrides: Dict[str, Any], dest: str,
                         rr_sheet: str = RR_SHEET) -> List[str]:
    """Fill BLANK unit cells of each overridden field with the analyst-supplied value, saving a
    gap-filled copy to `dest`. Only empty cells are written (never overwrites broker data).
    Returns audit notes. Values for kind=="pct" are divided by 100 (UI sends percentages)."""
    wb = openpyxl.load_workbook(rr_xlsx)
    ws = wb[rr_sheet]
    data_rows = _rr_data_rows(ws)
    notes: List[str] = []
    for field, raw in (overrides or {}).items():
        g = _GAP_BY_FIELD.get(field)
        if g is None or raw in (None, ""):
            continue
        try:
            val = float(raw)
        except (TypeError, ValueError):
            continue
        if g["kind"] == "pct":
            val = val / 100.0
        col = cix(g["column"])
        filled = 0
        for r in data_rows:
            if ws.cell(r, col).value in (None, ""):
                ws.cell(r, col).value = val
                filled += 1
        if filled:
            notes.append(f"gap override: filled {filled} blank {field} cell(s) with {raw}{g['unit']}")
    wb.save(dest)
    return notes


def _normalise(raw_path: str, asset: str, region: str, out_xlsx: str,
               mapping: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    client = None
    if mapping is None:
        try:
            import anthropic  # type: ignore
            client = anthropic.Anthropic()
        except Exception as e:
            raise HTTPException(status_code=500,
                                detail=f"Anthropic client init failed (ANTHROPIC_API_KEY?): {e}")
    try:
        return na.normalise_auto(raw_path, asset, region, out_xlsx,
                                 rr_sheet=RR_SHEET, mapping=mapping, anthropic_client=client)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Normalisation failed: {e}")


# ---------------------------------------------------------------- router factory

def make_underwrite_router(store) -> APIRouter:
    router = APIRouter(prefix="/underwrite", tags=["underwrite"])

    @router.post("/{deal_id}")
    async def underwrite_upload(
        deal_id: str,
        file: UploadFile = File(...),
        asset: Optional[str] = Form(None),
        region: Optional[str] = Form(None),
    ) -> Dict[str, Any]:
        """Upload a broker rent roll -> normalise (hand/known/LLM) -> Mode A flags (+rationale).
        Persists underwrite={status:"flagged",...}. Returns mapping + 3 sample rows + flags."""
        deal = _deal_or_404(store, deal_id)
        d = _defaults(deal)
        asset = asset or d["asset"]
        region = region or d["region"]

        ddir = _deal_dir(deal_id)
        raw_path = ddir / f"raw_{file.filename}"
        raw_path.write_bytes(await file.read())

        rr_xlsx = str(ddir / "canonical_rr.xlsx")
        norm = _normalise(str(raw_path), asset, region, rr_xlsx, mapping=None)
        mode_a = uw.run_mode_a(deal_id, rr_xlsx, RR_SHEET)
        flags = _enrich_flags([*(norm.get("flags") or []), *(mode_a.get("flags") or [])])

        block = _merge_underwrite(store, deal_id, {
            "status": "flagged",
            "asset": asset,
            "region": region,
            "raw_path": str(raw_path),
            "rr_xlsx": rr_xlsx,
            "rr_sheet": RR_SHEET,
            "mapping": norm["mapping"],
            "mapping_confirmed": bool(norm["mapping"].get("_source", "").startswith("hand")),
            "flags": flags,
            "schema_errors": mode_a["schema_errors"],
            "units": mode_a["units"],
            # audit scaffolding (untouched by upload; populated by /run)
            "baseline": (deal.get("underwrite") or {}).get("baseline"),
            "runs": (deal.get("underwrite") or {}).get("runs", []),
            "returns": None,
            "checks": None,
            "display_returns": False,
        })
        return {
            "deal_id": deal_id, "status": block["status"], "asset": asset, "region": region,
            "units": block["units"], "mapping": block["mapping"],
            "mapping_confirmed": block["mapping_confirmed"],
            "sample_rows": _sample_rows(rr_xlsx), "flags": block["flags"],
            "schema_errors": block["schema_errors"], "gaps": _compute_gaps(rr_xlsx),
        }

    @router.post("/{deal_id}/confirm-mapping")
    def underwrite_confirm_mapping(deal_id: str, body: ConfirmMappingBody) -> Dict[str, Any]:
        """Analyst-corrected column map -> re-run apply_mapping (deterministic, idempotent)."""
        deal = _deal_or_404(store, deal_id)
        block = deal.get("underwrite") or {}
        raw_path = block.get("raw_path")
        if not raw_path or not Path(raw_path).exists():
            raise HTTPException(status_code=409,
                                detail="No uploaded rent roll; POST /underwrite/{deal_id} first.")
        asset = body.asset or block.get("asset") or _defaults(deal)["asset"]
        region = body.region or block.get("region") or _defaults(deal)["region"]

        rr_xlsx = block.get("rr_xlsx") or str(_deal_dir(deal_id) / "canonical_rr.xlsx")
        norm = _normalise(raw_path, asset, region, rr_xlsx, mapping=body.mapping)
        mode_a = uw.run_mode_a(deal_id, rr_xlsx, RR_SHEET)
        flags = _enrich_flags([*(norm.get("flags") or []), *(mode_a.get("flags") or [])])

        # record the mapping correction in the audit log
        runs = list(block.get("runs") or [])
        runs.append({"version": None, "type": "mapping_correction", "at": _now(),
                     "analyst": body.analyst, "note": body.note or "manual column-map correction"})

        updated = _merge_underwrite(store, deal_id, {
            "status": "flagged", "asset": asset, "region": region, "rr_xlsx": rr_xlsx,
            "mapping": body.mapping, "mapping_confirmed": True, "flags": flags,
            "schema_errors": mode_a["schema_errors"], "units": mode_a["units"], "runs": runs,
        })
        return {
            "deal_id": deal_id, "status": updated["status"], "mapping_confirmed": True,
            "units": updated["units"], "mapping": updated["mapping"],
            "sample_rows": _sample_rows(rr_xlsx), "flags": updated["flags"],
            "schema_errors": updated["schema_errors"], "gaps": _compute_gaps(rr_xlsx),
        }

    @router.post("/{deal_id}/run")
    def underwrite_run(deal_id: str, body: RunBody) -> Dict[str, Any]:
        """Mode B: inject confirmed RR + assumptions, recalc, verify, return metrics. Appends a
        VERSIONED run (assumptions + diff vs baseline/previous + note). Refuses without sign-off."""
        deal = _deal_or_404(store, deal_id)
        block = deal.get("underwrite") or {}
        if block.get("status") not in ("flagged", "checks_failed", "underwritten"):
            raise HTTPException(status_code=409, detail="Run Mode A first (POST /underwrite/{deal_id}).")
        if not (body.mapping_signed_off and body.flags_signed_off):
            raise HTTPException(status_code=409, detail=(
                "Human-in-the-loop required: set mapping_signed_off and flags_signed_off to true "
                "once the column map and Mode A flags have been reviewed."))

        rr_xlsx = block.get("rr_xlsx")
        if not rr_xlsx or not Path(rr_xlsx).exists():
            raise HTTPException(status_code=409, detail="Canonical RR missing; re-upload the rent roll.")

        assumptions = dict(body.assumptions or {})
        entry = assumptions.get("entry_date")
        if not entry:
            raise HTTPException(status_code=422, detail="assumptions.entry_date (YYYY-MM-DD) is required.")

        asset = body.asset or block.get("asset") or _defaults(deal)["asset"]
        region = body.region or block.get("region") or _defaults(deal)["region"]

        run_dir = _deal_dir(deal_id) / f"run_{datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')}"
        run_dir.mkdir(parents=True, exist_ok=True)

        # Apply analyst gap overrides to a COPY of the canonical RR (blank cells only); the
        # original schedule is never mutated. Pass the filled copy to Mode B.
        gap_notes: List[str] = []
        rr_for_run = rr_xlsx
        if body.gap_overrides:
            filled = run_dir / "gapfilled_rr.xlsx"
            try:
                gap_notes = _apply_gap_overrides(rr_xlsx, body.gap_overrides, str(filled))
                if gap_notes:
                    rr_for_run = str(filled)
            except Exception as e:
                raise HTTPException(status_code=422, detail=f"Gap-override fill failed: {e}")

        try:
            result = uw.run_mode_b(deal_id, rr_for_run, RR_SHEET, asset, region, entry,
                                   assumptions=assumptions, workdir=str(run_dir))
        except Exception as e:
            import traceback as _tb
            _trace = _tb.format_exc()
            error_log_path = REPO_ROOT / "mode_b_error.log"
            try:
                error_log_path.write_text(_trace, encoding="utf-8")
                detail = f"Mode B failed: {e}\n\nFull traceback saved to {error_log_path}"
            except Exception:
                detail = f"Mode B failed: {e}\n\nFull traceback:\n{_trace}"
            print("\n=== Mode B traceback ===\n" + _trace + "\n=== end traceback ===\n", flush=True)
            raise HTTPException(status_code=500, detail=detail)

        model_dest = run_dir / "model.xlsx"
        try:
            shutil.copyfile(result["model_xlsx"], model_dest)
        except Exception:
            model_dest = Path(result["model_xlsx"])

        checks = result["checks"]
        passed = bool(checks.get("pass"))
        # Combine engine assumption notes with gap-override audit notes.
        all_assumption_notes = list(result.get("assumption_notes", [])) + gap_notes

        # ----- versioned audit record -----
        prior_runs = [r for r in (block.get("runs") or []) if r.get("type") != "mapping_correction"]
        prev_assumptions = prior_runs[-1]["assumptions"] if prior_runs else {}
        baseline = block.get("baseline")
        version = len(prior_runs) + 1
        run_record = {
            "version": version,
            "type": "run",
            "at": _now(),
            "analyst": body.analyst,
            "is_baseline": baseline is None,
            "assumptions": assumptions,
            "vs_previous": _assumptions_diff(prev_assumptions, assumptions),
            "vs_baseline": _assumptions_diff((baseline or {}).get("assumptions"), assumptions),
            "flag_resolutions": body.flag_resolutions or [],
            "gap_overrides": dict(body.gap_overrides or {}),
            "returns": result["returns"],
            "checks": checks,
            "assumption_notes": all_assumption_notes,
            "note": body.note or "",
            "model_xlsx": str(model_dest),
            "passed": passed,
        }
        all_runs = list(block.get("runs") or []) + [run_record]
        # freeze the machine baseline on the first run
        if baseline is None:
            baseline = {"at": run_record["at"], "assumptions": assumptions,
                        "returns": result["returns"], "checks": checks, "version": version}
        flags = _merge_flag_resolutions(block.get("flags") or [], body.flag_resolutions, body.analyst)

        updated = _merge_underwrite(store, deal_id, {
            "status": "underwritten" if passed else "checks_failed",
            "baseline": baseline,
            "runs": all_runs,
            "flags": flags,
            "assumptions": assumptions,
            "returns": result["returns"],
            "checks": checks,
            "assumption_notes": all_assumption_notes,
            "model_xlsx": str(model_dest),
            "display_returns": passed,
            "latest_version": version,
        })
        return {
            "deal_id": deal_id, "status": updated["status"], "version": version,
            "is_baseline": run_record["is_baseline"], "display_returns": passed,
            "returns": result["returns"], "checks": checks,
            "vs_baseline": run_record["vs_baseline"], "vs_previous": run_record["vs_previous"],
            "assumption_notes": all_assumption_notes,
            "model_url": f"/underwrite/{deal_id}/model",
        }

    @router.get("/{deal_id}")
    def underwrite_get(deal_id: str) -> Dict[str, Any]:
        deal = _deal_or_404(store, deal_id)
        block = deal.get("underwrite")
        if not block:
            raise HTTPException(status_code=404, detail=f"No underwrite for deal: {deal_id}")
        # Surface fresh per-field gap counts so the panel can render the gap-overrides section.
        out = dict(block)
        out["gaps"] = _compute_gaps(block.get("rr_xlsx"))
        return out

    @router.get("/{deal_id}/history")
    def underwrite_history(deal_id: str) -> Dict[str, Any]:
        """The full audit log: baseline + every versioned run with its diffs and notes."""
        deal = _deal_or_404(store, deal_id)
        block = deal.get("underwrite") or {}
        return {"deal_id": deal_id, "baseline": block.get("baseline"),
                "runs": block.get("runs", []), "flags": block.get("flags", [])}

    @router.get("/{deal_id}/model")
    def underwrite_model(deal_id: str, version: Optional[int] = None):
        """Download the populated, editable .xlsx (latest by default, or ?version=N)."""
        deal = _deal_or_404(store, deal_id)
        block = deal.get("underwrite") or {}
        path = block.get("model_xlsx")
        if version is not None:
            for r in block.get("runs", []):
                if r.get("version") == version and r.get("model_xlsx"):
                    path = r["model_xlsx"]
                    break
        if not path or not Path(path).exists():
            raise HTTPException(status_code=404, detail="No model for this deal/version yet.")
        asset = (block.get("asset") or deal_id).replace(" ", "_")
        tag = f"_v{version}" if version is not None else ""
        return FileResponse(
            path, filename=f"{asset}_underwrite{tag}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    return router
