#!/usr/bin/env python3
"""Post-recalc verification for an MLI underwrite run — v14/v21 (cross-check tool).

Usage: python verify.py <recalced.xlsx> [anchor_netcap=8203713]

Runs the v14 acceptance asserts + FEATURE checks (engine fixes, dynamic CFO, rates/SC,
YOC helpers) + v21 checks (Guarantee column inserted; ICS guarantee-income term removed).
Reads BOTH cached values (data_only=True) and formulas (data_only=False): the worst failure
mode — text in a SUMPRODUCT-consumed column — only shows as #VALUE! in Excel while LibreOffice
masks it as 0. See the v14 Build Spec.
v13/v21 structure: TI header row 28; West Craig 29-47; deal rows 48+; arrays 29-244;
template = TI - 5. WC tie-out target GA D16 = 8,203,713.
"""
import sys, openpyxl
from openpyxl.utils import column_index_from_string as cix
from openpyxl.worksheet.formula import ArrayFormula

TI_FIRST, TI_LAST = 29, 244
ANCHOR_DEFAULT = 8203713
NUMERIC_ARRAY_COLS = ['M','P','AA','U','AC','AD','AE','AX','AZ','BA','BB','BF','BO','BQ',
                      'BX','BY','BZ','CB','CF','CJ','CM','CN','CQ','CT','DA','DH','DI',
                      'DJ','DL','EI','EM','ER','EU','FV']

def v(ws,r,c): return ws.cell(r,cix(c)).value

def main():
    path=sys.argv[1]
    anchor=float(sys.argv[2]) if len(sys.argv)>2 else ANCHOR_DEFAULT
    wbv=openpyxl.load_workbook(path,data_only=True)
    wbf=openpyxl.load_workbook(path,data_only=False)
    tiv=wbv['Tenancy Inputs']; ga=wbv['Global Assumptions']; cfo=wbv['Cash Flow Output']; pcf=wbv['Project Cashflow']
    ok=True
    def check(label,passed,detail=''):
        nonlocal ok; ok&=bool(passed)
        print(f"[{'OK  ' if passed else 'FAIL'}] {label}"+((' -> '+detail) if detail and not passed else ''))
    rows=[r for r in range(TI_FIRST,TI_LAST+1)
          if isinstance(v(tiv,r,'C'),str) and v(tiv,r,'C').strip() and 'Total' not in v(tiv,r,'C')]
    # 1 anchor tie-out
    d16=ga['D16'].value
    check(f"West Craig tie-out GA D16 = {d16:,.2f} (target {anchor:,.0f})" if isinstance(d16,(int,float)) else "GA D16 non-numeric",
          isinstance(d16,(int,float)) and abs(d16-anchor)<1, str(d16))
    # 2 no zero exit on let units
    zero_exit=[r for r in rows if v(tiv,r,'CK') not in (None,'','Vacate')
               and isinstance(v(tiv,r,'EU'),(int,float)) and v(tiv,r,'EU')==0
               and isinstance(v(tiv,r,'AA'),(int,float)) and v(tiv,r,'AA')>0]  # exclude zero-ERV (substations/peppercorns)
    check("no let unit with zero exit value (EU)", not zero_exit, str(zero_exit))
    # 3 break/capex consistency
    tplf=wbf['TENANCY SCHEDULE (template)']
    n_ao=sum(1 for r in rows if v(tiv,r,'AO')=='T')
    n_break=sum(1 for r in range(24,240) if v(tplf,r,'M') in (1,'1'))
    print(f"[{'OK  ' if n_ao==n_break else 'WARN'}] AO='T' ({n_ao}) vs template breaks ({n_break})"
          + ("  (WARN: West Craig anchor breaks aren't in template M — expected when anchor toggled on)" if n_ao!=n_break else ""))
    n_capex=sum(1 for r in rows if isinstance(v(tiv,r,'CF'),(int,float)) and v(tiv,r,'CF')>0)
    print(f"[INFO] units with entry capex (CF>0): {n_capex}")
    # 4 error scan
    errs=[]
    for ws in wbv.worksheets:
        if ws.title=='TI Build Guide': continue
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value,str) and c.value in ('#REF!','#VALUE!','#NAME?','#DIV/0!','#NUM!'):
                    errs.append(f"{ws.title}!{c.coordinate}={c.value}")
    real=[e for e in errs if not e.startswith('BRE Economics!E75')]
    check("no error cells outside pre-existing BRE!E75", not real, str(real[:6]))
    # 5 Excel-safety: text in numeric/date array columns
    text_hits=[]
    for col in NUMERIC_ARRAY_COLS:
        for r in range(TI_FIRST,TI_LAST+1):
            cell=tiv.cell(r,cix(col)).value
            if isinstance(cell,str) and cell.strip()!='':
                text_hits.append(f"{col}{r}={cell!r}")
    check("no text in numeric/date array TI columns (Excel #VALUE trap)", not text_hits, str(text_hits[:8]))
    # 6 entry-date rewiring
    d5=ga['D5'].value; h9=pcf['H9'].value
    same=(hasattr(d5,'date') and hasattr(h9,'date') and d5.date()==h9.date())
    check("PCF purchase date (H9) == entry date (GA D5)", same, f"H9={h9} vs D5={d5}")
    f59=wbf['Cash Flow Output']['F59'].value; l20=wbf['Cash Flow Output']['L20'].value
    f59=f59.text if isinstance(f59,ArrayFormula) else f59
    l20=l20.text if isinstance(l20,ArrayFormula) else l20
    check("CFO row 59 stub factor dynamic (not hardcoded 4/3)", isinstance(f59,str) and 'MONTH(' in f59 and '4/3' not in f59, str(f59)[:50])
    check("CFO L20 uses hold length D8 (not /5)", isinstance(l20,str) and 'D$8' in l20, str(l20)[:50])
    # ---- v14 feature checks (engine fixes + dynamic CFO + rates/SC + YOC) ----
    tif=wbf['Tenancy Inputs']; cfof=wbf['Cash Flow Output']
    def txt(c):
        x=c.value; return x.text if isinstance(x,ArrayFormula) else x
    bre=wbf['BRE Economics'] if 'BRE Economics' in wbf.sheetnames else None
    check("FEATURE C13 NEF=1.0 (vacate re-let full ERV)", tif['C13'].value==1.0, str(tif['C13'].value))
    eu=tif['EU48'].value or ''
    check("FEATURE EU exit deduction = void rent only (no +I11)", '$I$11' not in eu and '$I$10/12' in eu, eu[:50])
    if bre is not None:
        ab2=txt(bre['AB2']) or ''
        check("FEATURE promote AB2 dynamic for any hold", 'MATCH' in ab2 and 'IF(INDEX' in ab2, ab2[:40])
    p5=txt(cfof['P5'])
    check("FEATURE register P5 dynamic (auto from TI col C)", isinstance(p5,str) and 'MATCH(0' in p5, str(p5)[:40])
    x17=txt(cfof['X17']) or ''; y17=txt(cfof['Y17']) or ''
    check("FEATURE YOC untrended uses FX+FZ; trended uses FY", 'FX' in x17 and 'FZ' in x17 and 'FY' in y17, "X/Y formulas")
    check("FEATURE TI helper cols FY/FX/FZ present", all(tif[f'{c}48'].value not in (None,'') for c in ['FY','FX','FZ']), "FY/FX/FZ")
    ax=tif['AX48'].value or ''; bb=tif['BB48'].value or ''
    check("FEATURE rates/SC wired (TI AX<-tplW, BB<-tplX)", 'SCHEDULE' in ax and 'SCHEDULE' in bb, f"{ax[:20]}|{bb[:20]}")
    ab=tif['AB48'].value or ''
    check("FEATURE vacant@entry capitalises guarantee/headline U (AB uses $U)", '$U48' in ab, ab[:45])
    bf=tif['BF48'].value or ''
    check("FEATURE break units skip rent review (BF has AO<>'T')", '$AO48<>"T"' in bf, bf[:45])
    du=tif['DU48'].value or ''
    check("FEATURE vacant@entry renews at Lease 3 (DU not cascading Vacate)", '$S48="Y","Renewal"' in du, du[:45])
    # ---- v21 checks: Guarantee column inserted; ICS legacy guarantee-income term removed ----
    import re as _re
    GUARANTEE_HDR='Guarantee Rent ('+chr(0xA3)+'pa)'   # avoid a literal pound byte in source
    hdrs={(tplf.cell(17,c).value or '').strip() for c in range(1,tplf.max_column+1)}
    check("v21 Guarantee Rent column present in template (header row 17)",
          any(h.startswith('Guarantee Rent') for h in hdrs), "missing header")
    u_text=[f"U{r}" for r in range(48,TI_LAST+1)
            if isinstance(v(tiv,r,'U'),str) and v(tiv,r,'U').strip()!='']
    check("v21 TI U numeric (Guarantee col is 0 not text empty-string)", not u_text, str(u_text[:6]))
    icsf=wbf['Income & Cost Schedule']
    gpat=_re.compile(r"\('Tenancy Inputs'!\$T\d+=\"Y\"\)\*.*?'Tenancy Inputs'!\$U\d+")
    def _t(c):
        x=c.value; return x.text if isinstance(x,ArrayFormula) else x
    gterm=sum(1 for row in icsf.iter_rows() for c in row
              if isinstance(_t(c),str) and gpat.search(_t(c)))
    check("v21 ICS free of legacy guarantee-income term (T=Y * U)", gterm==0, f"{gterm} cells still carry it")
    print("\nHeadline: Unlevered IRR", cfo['L17'].value, "| Net Investor", cfo['K18'].value, "| EM", cfo['J19'].value)
    print("REMINDER: open in Excel (Ctrl+Alt+F9), eyeball headline + Next Buyer capex.")
    print("\nRESULT:", "PASS" if ok else "FAIL — investigate above")
    sys.exit(0 if ok else 1)

if __name__=='__main__':
    main()
