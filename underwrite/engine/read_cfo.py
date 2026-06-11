"""
CFO returns extraction — label-resolved cell lookup for v21 models.

Reads a recalculated model and extracts the unlevered / levered (geared) returns
block by locating the KPI matrix on the Cash Flow Output sheet by its row labels
and column headers — not hardcoded cell letters. Returns None if the matrix cannot
be resolved (flag-don't-fabricate).

v21 KPI matrix shape (Cash Flow Output):

              J(Investment)  K(Investor PTF&Promote)  L(Unlevered)  M(Unlevered PTF)
  I  Equity        ...              ...                   ...            ...
  I  Unlevered IRR n/a              n/a                   <value>        ...
  I  Levered IRR   <value>          ...                   n/a            ...
  I  Equity mult.  ...              ...                   ...            ...
  I  Cash-on-cash  ...              n/a                   ...            ...
  I  Profit        ...              ...                   ...            ...

Metrics are ROW labels (column I); investment views are COLUMNS. The unlevered
view is the column headed exactly "Unlevered"; the levered/geared view is the
column headed "Investment" (where the Levered IRR row carries a value).
"""
from pathlib import Path
from typing import Dict, Any, Optional, Tuple
import openpyxl
from openpyxl.utils import get_column_letter


def extract_cfo_returns(recalced: Path) -> Optional[Dict[str, Any]]:
    """
    Extract CFO returns and assumptions from a recalculated v21 model.

    Args:
        recalced: Path to the recalc_out .xlsx file (data_only=True values).

    Returns:
        Dict with unlevered/levered returns, assumptions, and source traceability.
        None if the KPI matrix cannot be resolved.
    """
    try:
        wb = openpyxl.load_workbook(recalced, data_only=True)
    except Exception:
        return None

    # Find CFO sheet (case-insensitive)
    cfo_sheet = None
    for name in wb.sheetnames:
        if name.lower() == "cash flow output":
            cfo_sheet = wb[name]
            break
    if not cfo_sheet:
        return None

    # Resolve the view columns (unlevered / levered) + label column from the header row.
    unlev_col, lev_col, label_col, header_row = _find_view_columns(cfo_sheet)
    if unlev_col is None or lev_col is None or label_col is None:
        return None

    # Resolve the metric rows by their KPI-column labels, around the header row.
    rows = _find_metric_rows(cfo_sheet, header_row, label_col)
    required = ["unlev_irr", "lev_irr", "em", "coc", "profit", "equity"]
    if not all(k in rows for k in required):
        return None

    def cell(r: int, c: int) -> Optional[float]:
        return _safe_float(cfo_sheet.cell(r, c).value)

    unlev_returns = {
        "irr": cell(rows["unlev_irr"], unlev_col),
        "em": cell(rows["em"], unlev_col),
        "coc_y5": cell(rows["coc"], unlev_col),
        "profit": cell(rows["profit"], unlev_col),
        "equity": cell(rows["equity"], unlev_col),
    }
    lev_returns = {
        "irr": cell(rows["lev_irr"], lev_col),
        "em": cell(rows["em"], lev_col),
        "coc_y5": cell(rows["coc"], lev_col),
        "profit": cell(rows["profit"], lev_col),
        "equity": cell(rows["equity"], lev_col),
    }

    # The unlevered IRR is the load-bearing value; if it didn't resolve to a
    # number the matrix mapping is wrong — bail rather than emit a hollow block.
    if unlev_returns["irr"] is None or lev_returns["irr"] is None:
        return None

    assumptions = _extract_assumptions(wb)

    source = {
        "sheet": "Cash Flow Output",
        "unlevered_col": get_column_letter(unlev_col),
        "levered_col": get_column_letter(lev_col),
        "resolved_by": "label",
        "row_map": dict(rows),
    }

    return {
        "scenario_label": None,
        "unlevered": unlev_returns,
        "levered": lev_returns,
        "assumptions_table": assumptions,
        "source": source,
    }


def _cell_text(cell_sheet, row: int, col: int) -> str:
    """Lower-cased string form of a cell value (numbers/dates coerced safely)."""
    v = cell_sheet.cell(row, col).value
    if v is None:
        return ""
    return str(v).strip().lower()


def _find_view_columns(
    cfo_sheet,
) -> Tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
    """
    Scan top rows for the KPI header row and resolve the unlevered + levered
    (geared) view columns plus the label column.

    - unlevered view  = column whose header is exactly "unlevered"
    - levered view    = column whose header is "investment" (the geared view)
    - label column    = column whose header is "kpi" (holds the metric row labels);
                        falls back to the column immediately left of the views.

    Returns (unlev_col, lev_col, label_col, header_row) or all-None.
    """
    max_col = min(cfo_sheet.max_column, 30)
    for row_num in range(1, min(26, cfo_sheet.max_row + 1)):
        headers = {c: _cell_text(cfo_sheet, row_num, c) for c in range(1, max_col + 1)}

        unlev_col = next((c for c, t in headers.items() if t == "unlevered"), None)
        lev_col = next((c for c, t in headers.items() if t == "investment"), None)
        if unlev_col and lev_col:
            label_col = next((c for c, t in headers.items() if t == "kpi"), None)
            if label_col is None:
                label_col = min(unlev_col, lev_col) - 1
            return unlev_col, lev_col, label_col, row_num

    return None, None, None, None


def _find_metric_rows(cfo_sheet, header_row: int, label_col: int) -> Dict[str, int]:
    """
    Scan the KPI label column around the header row for the metric row labels.

    Anchoring to the single label column avoids colliding with the unrelated
    debt-assumptions table that shares these rows in columns C-F.
    """
    rows: Dict[str, int] = {}

    # KPIs can sit just above the header row (e.g. Equity) or below it; scan a
    # window straddling the header row.
    for row_num in range(max(1, header_row - 2), min(header_row + 20, cfo_sheet.max_row + 1)):
        label = _cell_text(cfo_sheet, row_num, label_col)
        if not label:
            continue

        if "unlevered irr" in label and "unlev_irr" not in rows:
            rows["unlev_irr"] = row_num
        elif "levered irr" in label and "lev_irr" not in rows:
            rows["lev_irr"] = row_num
        elif "equity multiple" in label and "em" not in rows:
            rows["em"] = row_num
        elif ("cash-on-cash" in label or "cash on cash" in label) and "coc" not in rows:
            rows["coc"] = row_num
        elif "profit" in label and "profit" not in rows:
            rows["profit"] = row_num
        elif label == "equity" and "equity" not in rows:
            rows["equity"] = row_num

    return rows


def _extract_assumptions(wb) -> Dict[str, Any]:
    """
    Extract a small assumptions table from the Global Assumptions sheet.

    Currently surfaces the net purchase price (GA!D16) which the engine already
    treats as the anchor; other fields are left None (honest TBC) until a stable
    GA label-scan is added.
    """
    ga = None
    for name in wb.sheetnames:
        if name.lower() == "global assumptions":
            ga = wb[name]
            break

    assumptions: Dict[str, Any] = {
        "net_purchase_price": None,
    }
    if ga is not None:
        assumptions["net_purchase_price"] = _safe_float(ga["D16"].value)
    return assumptions


def _safe_float(val: Any) -> Optional[float]:
    """Convert a value to float, or None if invalid (e.g. 'n/a')."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None
