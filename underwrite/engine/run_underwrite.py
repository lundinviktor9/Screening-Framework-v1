#!/usr/bin/env python3
"""MLI underwrite — Mode B compute & extract (v13). Recalcs headless; ALWAYS open the result in Excel
(Ctrl+Alt+F9) to confirm — a text-in-array #VALUE only shows in Excel (LibreOffice masks it).

Recalculates a populated model via LibreOffice (no Excel needed), runs the acceptance
asserts, and extracts the full return set to returns.json keyed by deal_id.

Usage:
  python run_underwrite.py <model.xlsx> --deal-id NEWBURY [--anchor-netcap 8203713]
                           [--out returns.json] [--recalc-out recalced.xlsx]

Requires: soffice (LibreOffice), python-uno, openpyxl. Assumes the template is already
populated and the cascade rewired (by mli-underwrite-runner). The anchor (West Craig)
should be toggled OFF for the deal headline; pass --anchor-netcap to also verify the
anchor ties out (toggle it on in a separate pass).
"""
import sys, os, json, argparse, subprocess, time

def recalc(inp, outp, port=2002):
    import uno
    from com.sun.star.beans import PropertyValue
    proc = subprocess.Popen(['soffice','--headless','--invisible','--nologo','--norestore',
        f'--accept=socket,host=localhost,port={port};urp;'],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    ctx=None; lc=uno.getComponentContext()
    res=lc.ServiceManager.createInstanceWithContext("com.sun.star.bridge.UnoUrlResolver",lc)
    for _ in range(60):
        try:
            ctx=res.resolve(f"uno:socket,host=localhost,port={port};urp;StarOffice.ComponentContext"); break
        except Exception: time.sleep(1)
    if not ctx: proc.kill(); raise RuntimeError("LibreOffice did not start")
    smgr=ctx.ServiceManager
    desktop=smgr.createInstanceWithContext("com.sun.star.frame.Desktop",ctx)
    def p(n,v):
        pv=PropertyValue(); pv.Name=n; pv.Value=v; return pv
    doc=desktop.loadComponentFromURL("file://"+os.path.abspath(inp),"_blank",0,(p("Hidden",True),))
    doc.calculateAll()
    doc.storeToURL("file://"+os.path.abspath(outp),(p("FilterName","Calc MS Excel 2007 XML"),))
    doc.close(False); desktop.terminate(); proc.wait(timeout=20)

def extract(path):
    import openpyxl
    from openpyxl.utils import column_index_from_string as cix
    wb=openpyxl.load_workbook(path, data_only=True)
    cfo=wb['Cash Flow Output']; ga=wb['Global Assumptions']; ti=wb['Tenancy Inputs']
    def num(x): return x if isinstance(x,(int,float)) else None
    out={
      'net_purchase_price': num(ga['D16'].value),
      'unlevered_irr':  num(cfo['L17'].value),
      'net_investor_irr': num(cfo['K18'].value),
      'levered_irr_pretax_J18': num(cfo['J18'].value),
      'equity_multiple_levered_J19': num(cfo['J19'].value),
      'equity_multiple_investor_K19': num(cfo['K19'].value),
      'equity_multiple_unlevered_L19': num(cfo['L19'].value),
      'cash_on_cash_J20': num(cfo['J20'].value),
    }
    if 'Next Buyer CF' in wb.sheetnames:
        nb=wb['Next Buyer CF']
        out['next_buyer_cost_basis_m']=num(nb['C4'].value)
        out['next_buyer_exit_price_m']=num(nb['C5'].value)
    # checks
    rows=[r for r in range(29,245)
          if isinstance(ti.cell(r,cix('C')).value,str) and ti.cell(r,cix('C')).value.strip()
          and 'Total' not in ti.cell(r,cix('C')).value]
    zero_exit=[r for r in rows if ti.cell(r,cix('CK')).value not in (None,'','Vacate')
               and num(ti.cell(r,cix('EU')).value)==0]
    errs=sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row
             if isinstance(c.value,str) and c.value.startswith('#') and c.value.endswith(('!','?')))
    out['_checks']={'let_units_zero_exit': zero_exit, 'workbook_error_cells': errs,
                    'pass': (not zero_exit and errs==0)}
    return out

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument('model'); ap.add_argument('--deal-id', default=None)
    ap.add_argument('--anchor-netcap', type=float, default=None)
    ap.add_argument('--out', default='returns.json')
    ap.add_argument('--recalc-out', default=None)
    a=ap.parse_args()
    deal=a.deal_id or os.path.splitext(os.path.basename(a.model))[0]
    recalced=a.recalc_out or (os.path.splitext(a.model)[0]+'_recalc.xlsx')
    recalc(a.model, recalced)
    res=extract(recalced)
    res={'deal_id': deal, **res}
    if a.anchor_netcap is not None:
        import openpyxl
        d16=openpyxl.load_workbook(recalced, data_only=True)['Global Assumptions']['D16'].value
        res['_checks']['anchor_tieout_ok']=abs(d16 - a.anchor_netcap) < 1
    json.dump(res, open(a.out,'w'), indent=2)
    print(json.dumps(res, indent=2))
    print('\nPASS' if res['_checks'].get('pass') else '\nCHECK FAILED — review _checks')
    sys.exit(0 if res['_checks'].get('pass') else 1)

if __name__=='__main__':
    main()
