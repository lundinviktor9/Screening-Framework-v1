#!/usr/bin/env python3
"""inject_deal_v21.py - inject a normalised RR into a v21-layout clean base + wire it.

Refit of inject_deal_v14 for the v20/v21 template layout (Guarantee Rent column inserted
between Passing and ERV; every template column >= T shifted +1). Key changes:

- TEMPLATE columns are resolved BY HEADER (row 17), not hardcoded letters, so the injector
  RE-DERIVES from the live model and survives column inserts/moves. TI's own columns did NOT
  move in v20/v21, so TI letters stay hardcoded (stable).
- New "Guarantee Rent (Lpa)" column (vacant-only PP input) written NUMERIC:
  =IF($P=Y, RR!BF, 0) with a hidden-zero number format - never "" (the v21 Excel-safety fix;
  a text "" reaching an array formula is the #VALUE! trap).
- The old far-right BF guarantee-rent write is REMOVED (TI U now reads the new column).
- TI per-unit cascade is filled for ALL deal rows from a clean SRC row, so per-unit formulas
  track whatever the live model carries; explicit overrides are kept for the genuinely
  per-unit/data columns and reference header-resolved template void/RF columns.

RR sheet follows the field-dictionary OLD letter scheme; FIELD_MAP maps RR letter -> v21
template HEADER. Validate by WC tie-out (GA D16 = 8,203,713.29) + ICS error-free + consistency,
NOT penny-reproduction of a hand-tweaked deal (see spec S14.3).

Usage: inject_deal_v21.py --base B --rr-sheet S --asset A --region R --entry YYYY-MM-DD --out O [--src-row 60]
"""
import argparse, datetime, re, copy, openpyxl
from openpyxl.utils import column_index_from_string as cix, get_column_letter as gl
TPL = "TENANCY SCHEDULE (template)"
TPL_FIRST = 43; TI_FIRST = 48; OFF = 5; CASCADE_SRC_DEFAULT = 60

# RR (field-dictionary OLD letter)  ->  v21 template HEADER (exact, row 17)
FIELD_MAP = [
    ('A', '#'), ('B', 'Asset Name'), ('C', 'Region'), ('D', 'Sector'), ('E', 'Unit Number'),
    ('F', 'Tenant Name'), ('G', 'Area GIA (sq ft)'), ('H', 'Entry Yield (NIY)'),
    ('J', 'Lease Start'), ('K', 'Lease Expiry'), ('L', 'Break Date'),
    ('M', 'Break Taken (1=Yes,0=No)'), ('N', 'Rent Review / MTM Date'),
    ('O', 'Event @ Expiry (Y=Renew / X=Vacate)'), ('P', 'Vacant @ Entry (Y/N)'),
    ('Q', 'Rent Review NEF'), ('R', 'ERV Growth to Lease Start (% pa)'),
    ('S', 'Passing Rent (£ pa)'),
    ('T', 'ERV (£ pa)'), ('U', 'ERV (£ psf) [optional]'),
    ('V', 'Rateable Value (£)'), ('W', 'Business Rates (£ pa)'), ('X', 'Service Charge (£ pa)'),
    ('Y', 'Assumed Void (mths)'), ('Z', 'Assumed Rent Free (mths)'), ('AA', 'Re-letting Capex (£ psf)'),
    ('AY', '1954 Act (Y/N)'), ('AZ', 'EPC Rating'), ('BB', 'Rent Review (Y/N)'),
    ('BC', 'Term Certain (mths)'), ('BG', 'Guarantee Period (mths)'),
]
GUARANTEE_HDR = 'Guarantee Rent (£pa)'   # new vacant-only PP column (no space before pa); fed from RR!BF
EXITYIELD_HDR = 'Exit Yield'
VACANT_HDR = 'Vacant @ Entry (Y/N)'
VOID_HDR = 'Assumed Void (mths)'
RF_HDR = 'Assumed Rent Free (mths)'

_tok = re.compile(r"(?:'([^']+)'!)?(\$?)([A-Z]{1,3})(\$?)(\d+)")
def shift(formula, delta):
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula
    def rep(m):
        sheet, dc, col, dr, row = m.group(1), m.group(2), m.group(3), m.group(4), int(m.group(5))
        nr = row
        if sheet is None:
            if dr != '$' and 29 <= row <= 244:
                nr = row + delta
        elif sheet == TPL:
            if 24 <= row <= 180:
                nr = row + delta
        return (f"'{sheet}'!" if sheet is not None else "") + f"{dc}{col}{dr}{nr}"
    return _tok.sub(rep, formula)

def header_map(tpl, hrow=17):
    H = {}
    for c in range(1, tpl.max_column + 1):
        v = tpl.cell(hrow, c).value
        if isinstance(v, str) and v.strip():
            H.setdefault(v.strip(), gl(c))   # first occurrence wins (new T before far-right legacy)
    return H

def main():
    ap = argparse.ArgumentParser()
    for o in ('base', 'rr-sheet', 'asset', 'region', 'entry', 'out'):
        ap.add_argument('--' + o, required=True)
    ap.add_argument('--src-row', type=int, default=CASCADE_SRC_DEFAULT)
    a = ap.parse_args()
    wb = openpyxl.load_workbook(a.base)
    rrname = getattr(a, 'rr_sheet')
    rr = wb[rrname]; tpl = wb[TPL]; ti = wb['Tenancy Inputs']
    cfo = wb['Cash Flow Output']; ga = wb['Global Assumptions']; pcf = wb['Project Cashflow']
    rs = f"'{rrname}'"
    H = header_map(tpl)
    def tcol(hdr):
        if hdr not in H:
            raise SystemExit(f"template header not found: {hdr!r}")
        return H[hdr]
    n = sum(1 for r in range(2, rr.max_row + 1) if rr.cell(r, cix('E')).value)
    TI_LAST = TI_FIRST + n - 1; TPL_LAST = TPL_FIRST + n - 1; SRC = a.src_row; maxc = ti.max_column
    vac_col = tcol(VACANT_HDR); void_col = tcol(VOID_HDR); rf_col = tcol(RF_HDR)
    gcol = tcol(GUARANTEE_HDR)
    print(f"units={n}  TI {TI_FIRST}-{TI_LAST}  template {TPL_FIRST}-{TPL_LAST}  "
          f"guarantee->{gcol}  void->{void_col}  rf->{rf_col}")
    # 0. fill TI cascade for ALL deal rows from SRC (carries live v21 per-unit formulas)
    for r in range(TI_FIRST, TI_LAST + 1):
        if r == SRC:
            continue
        d = r - SRC
        for c in range(1, maxc + 1):
            src = ti.cell(SRC, c).value
            ti.cell(r, c).value = shift(src, d) if (isinstance(src, str) and src.startswith('=')) else src
    # 1. template deal rows = RR refs (header-resolved) + new numeric guarantee column
    for i in range(n):
        r = TPL_FIRST + i; rrow = 2 + i
        for rr_letter, hdr in FIELD_MAP:
            tpl.cell(r, cix(tcol(hdr))).value = f"={rs}!{rr_letter}{rrow}"
        tpl.cell(r, cix(tcol(EXITYIELD_HDR))).value = f"={rs}!I{rrow}+'Cash Flow Output'!$Q$27"
        tpl.cell(r, cix(gcol)).value = f"=IF(${vac_col}{r}=\"Y\",{rs}!BF{rrow},0)"  # vacant-only, NUMERIC 0
        tpl.cell(r, cix(gcol)).number_format = '[=0]"-";#,##0'
    # formatting: copy a populated row's style; empty dates show "-"; re-apply guarantee fmt
    REFROW = 24
    for i in range(n):
        r = TPL_FIRST + i
        for c in range(1, tpl.max_column + 1):
            ref = tpl.cell(REFROW, c); cell = tpl.cell(r, c)
            cell.font = copy.copy(ref.font); cell.border = copy.copy(ref.border)
            cell.fill = copy.copy(ref.fill); cell.alignment = copy.copy(ref.alignment)
            cell.number_format = ref.number_format
        for dh in ('Lease Start', 'Lease Expiry', 'Break Date', 'Rent Review / MTM Date'):
            tpl.cell(r, cix(tcol(dh))).number_format = '[=0]"-";dd/mm/yyyy'
        tpl.cell(r, cix(gcol)).number_format = '[=0]"-";#,##0'
    for r in range(TPL_LAST + 1, TPL_FIRST + 135):
        for c in range(1, tpl.max_column + 1):
            tpl.cell(r, c).value = None
    # 2. per-unit overrides (TI cols stable; template void/RF header-resolved) + per-unit data
    for i in range(n):
        r = TI_FIRST + i; tr = r - OFF; rrow = 2 + i
        ti.cell(r, cix('C')).value = rr.cell(rrow, cix('B')).value   # per-row asset name
        ti.cell(r, cix('E')).value = rr.cell(rrow, cix('C')).value   # per-row region
        ti.cell(r, cix('AB')).value = f'=IF($S{r}="Y",$U{r},$P{r})'
        ti.cell(r, cix('AD')).value = (f'=IF($S{r}="Y",$U{r}*($I$10/12)+$AA{r}*$I$11'
                                       f'+$AX{r}*MAX(0,$I$8-$L$8)/12+$BB{r}*($I$8/12),0)')
        ti.cell(r, cix('CL')).value = f"=IF($S{r}=\"Y\",$I$8,'{TPL}'!${void_col}{tr})"
        ti.cell(r, cix('CN')).value = f"=IF($S{r}=\"Y\",$I$9,'{TPL}'!${rf_col}{tr})"
        ti.cell(r, cix('AF')).value = 0; ti.cell(r, cix('BY')).value = 1
        ti.cell(r, cix('EU')).value = (f'=IF($ET{r}=0,0,IF($EK{r}="Y",$ER{r}/$ET{r}'
                                       f'-$ER{r}*($I$10/12),$EM{r}/$ET{r}))')
        ti.cell(r, cix('BF')).value = (f'=IF(AND($AM{r}="Y",$M{r}<50000,$AO{r}<>"T"),'
                                       f'IFERROR(($AA{r}*$BE{r})*(1+$BD{r})^(YEARFRAC($C$6,$M{r})),$AA{r}),0)')
        ti.cell(r, cix('EM')).value = (f'=IF($EI{r}<=$CJ{r},IF(AND($AM{r}="Y",$M{r}<50000,$EI{r}>=$M{r},$AO{r}<>"T"),'
                                       f'$BF{r},$P{r}),IF($EI{r}<$CM{r},0,$CQ{r}))')
        ti.cell(r, cix('DU')).value = f'=IF($S{r}="Y","Renewal",$CK{r})'
        cx = rr.cell(rrow, cix('AA')).value
        ti.cell(r, cix('BW')).value = ('Y' if isinstance(cx, (int, float)) and cx > 0 else 'N')
    for r in range(TI_LAST + 1, TI_FIRST + 135):
        ti.cell(r, cix('C')).value = None; ti.cell(r, cix('E')).value = None
    # 3. register + 4. entry date + 5. entry-date rewiring
    cfo['Q6'] = 1; cfo['Q5'] = 0   # register P5:P10 auto-detects asset names from TI col C
    y, m, d = map(int, a.entry.split('-')); ga['D5'] = datetime.datetime(y, m, d)
    pcf['H8'] = "='Global Assumptions'!$D$5"; pcf['H9'] = "='Global Assumptions'!$D$5"
    pcf['J8'] = "=DATE(YEAR($I$8),MONTH($I$8)+1,1)"
    GA = "'Global Assumptions'"
    for c in range(cix('F'), cix('L') + 1):
        L = gl(c)
        cfo.cell(59, c).value = (f"=IF({L}25=YEAR({GA}!$D$5),12/(12-MONTH({GA}!$D$5)),"
                                 f"IF({L}25=YEAR({GA}!$D$9),12/MONTH({GA}!$D$9),"
                                 f"IF(AND({L}25>YEAR({GA}!$D$5),{L}25<YEAR({GA}!$D$9)),1,0)))")
    cfo['L20'] = f"=IFERROR(SUM(F40:L40)/({GA}!$D$8/12)/(-E44),\"n/a\")"
    cfo['M20'] = "=IFERROR(AVERAGE(F63:J63),\"n/a\")"
    wb.save(a.out)
    print("wrote", a.out)

if __name__ == '__main__':
    main()
