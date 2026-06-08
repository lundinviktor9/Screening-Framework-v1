#!/usr/bin/env python3
"""Meadow layout -> normalised RR (one estate at a time via --asset-filter).
Wires rates (AE->W) and service charge (AF->X). Per-asset yields from schedule."""
import openpyxl, sys, datetime as dt, argparse
from openpyxl.utils import column_index_from_string as cix
def pdate(x):
    if isinstance(x,dt.datetime): return x
    if isinstance(x,dt.date): return dt.datetime(x.year,x.month,x.day)
    if isinstance(x,str):
        for f in ('%d/%m/%Y','%d/%m/%y','%Y-%m-%d'):
            try: return dt.datetime.strptime(x.strip(),f)
            except: pass
    return None
def mny(x): return float(x) if isinstance(x,(int,float)) else None
ap=argparse.ArgumentParser()
ap.add_argument('--src',required=True); ap.add_argument('--asset-filter',required=True)
ap.add_argument('--region',required=True); ap.add_argument('--out',required=True)
a=ap.parse_args()
ws=openpyxl.load_workbook(a.src,data_only=True)["Meadow "]
M=dict(asset='A',region='B',unit='C',tenant='D',gia='E',eyld='G',xyld='H',start='L',expiry='M',
       brkdate='O',brktaken='P',review='Q',event='AJ',passing='W',ervpa='AA',ervpsf='AC',
       rv='AD',rates='AE',sc='AF',void='AU',rf='AV',capex='AR')
def g(r,k): return ws.cell(r,cix(M[k])).value
hdrmap={'A':'#','B':'Asset Name','C':'Region','D':'Sector','E':'Unit Number','F':'Tenant Name',
  'G':'Area GIA (sq ft)','H':'Entry Yield (NIY)','I':'Exit Yield','J':'Lease Start','K':'Lease Expiry',
  'L':'Break Date','M':'Break Taken (1=Yes,0=No)','N':'Rent Review / MTM Date',
  'O':'Event @ Expiry (Y=Renew / X=Vacate)','P':'Vacant @ Entry (Y/N)','S':'Passing Rent (£ pa)',
  'T':'ERV (£ pa)','U':'ERV (£ psf)','V':'Rateable Value','W':'Business Rates (£ pa)','X':'Service Charge (£ pa)',
  'Y':'Assumed Void (mths)','Z':'Assumed Rent Free (mths)','AA':'Re-letting Capex (£ psf)',
  'BB':'Rent Review (Y/N)','BC':'Term Certain (mths)','BE':'Rental Guarantee (Y/N)','BF':'Guarantee/Deduction Rent (£ pa)','BG':'Guarantee Period (mths)'}
out=openpyxl.Workbook(); o=out.active; o.title='RR'
for c,l in hdrmap.items(): o.cell(1,cix(c)).value=l
n=0; flags=[]
for r in range(2,ws.max_row+1):
    asset=g(r,'asset')
    if not asset or 'Total' in str(asset) or a.asset_filter.lower() not in str(asset).lower(): continue
    n+=1; orow=1+n
    tenant=str(g(r,'tenant') or ''); vacant='vacant' in tenant.lower()
    brk = (g(r,'brktaken') in (1,'1')) and not ('vacant' in tenant.lower())
    event=str(g(r,'event') or 'Y').strip().upper()
    if vacant: event='X'
    area=mny(g(r,'gia')); ervpsf=mny(g(r,'ervpsf')); ervpa=mny(g(r,'ervpa'))
    relet=(event=='X') and not vacant
    Nrev=pdate(g(r,'review'))
    vals={'A':n,'B':asset,'C':a.region,'D':'Industrial','E':g(r,'unit'),'F':tenant,'G':area,
      'H':mny(g(r,'eyld')),'I':mny(g(r,'xyld')),
      'J':None if vacant else pdate(g(r,'start')),'K':None if vacant else pdate(g(r,'expiry')),
      'L':pdate(g(r,'brkdate')) if brk else None,'M':1 if brk else 0,
      'N':Nrev if (Nrev and not vacant) else None,'O':event,'P':'Y' if vacant else 'N',
      'S':0 if vacant else mny(g(r,'passing')),'T':ervpa,'U':ervpsf,
      'V':mny(g(r,'rv')),'W':mny(g(r,'rates')),'X':mny(g(r,'sc')),
      'Y':((mny(g(r,'void')) if mny(g(r,'void')) is not None else 9) if relet else None),
      'Z':((mny(g(r,'rf')) if mny(g(r,'rf')) is not None else 6) if relet else None),
      'AA':((mny(g(r,'capex')) if mny(g(r,'capex')) is not None else 20) if relet else None),
      'BB':'Y' if (Nrev and not vacant) else 'N','BC':60,
      'BE':'Y' if vacant else 'N','BF':ervpa if vacant else None,'BG':12 if vacant else None}
    for c,v in vals.items(): o.cell(orow,cix(c)).value=v
    if vacant: flags.append(f"U{n} {vals['E']}: VACANT@entry -> ERV cap £{ervpa:,.0f}, re-let I8/I9")
    if brk: flags.append(f"U{n} {vals['E']}: BREAK taken -> vacate+re-let (void {vals['Y']}/RF {vals['Z']}/capex {vals['AA']})")
    if event=='X' and not brk and not vacant: flags.append(f"U{n} {vals['E']}: VACATE@expiry -> re-let")
out.save(a.out)
print(f"Normalised {n} units ({a.asset_filter}) -> {a.out}")
print(f"FLAGS: {len(flags)}")
for f in flags: print("  -",f)
