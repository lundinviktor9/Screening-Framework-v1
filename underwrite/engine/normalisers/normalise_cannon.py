#!/usr/bin/env python3
"""Cannon Ind. -> normalised 'Cannon RR' (template-letter cols A..BG). v14 shakedown.
Corrections vs first pass: per-unit capex (25 only on break/vacate let units, else 0);
vacant units carry T/AA=ERV (ERV-capitalisation per Viktor) and BF/U=guarantee rent
(deducted off PP); under-offer dated to entry +5yr."""
import openpyxl, datetime as dt, sys, re
from openpyxl.utils import column_index_from_string as cix

SRC=sys.argv[1]; OUT=sys.argv[2]
ENTRY=dt.datetime(2026,6,30); HOLD_MTHS=60
ASSET="Cannon Industrial Estate, Milton Keynes"; REGION="Milton Keynes"
EXITDATE=dt.datetime(2031,6,30)

def pdate(x):
    if isinstance(x,dt.datetime): return x
    if isinstance(x,dt.date): return dt.datetime(x.year,x.month,x.day)
    if isinstance(x,str):
        for f in ('%d/%m/%Y','%d/%m/%y','%Y-%m-%d'):
            try: return dt.datetime.strptime(x.strip(),f)
            except: pass
    return None
def mny(x): return float(x) if isinstance(x,(int,float)) else None

wb=openpyxl.load_workbook(SRC,data_only=True); ws=wb["Cannon Ind."]
C=dict(unit=1,tenant=2,gia=3,start=4,brkdate=5,review=6,expiry=7,brk=8,vacating=9,
        mrent_pa=12,lvoid=16,rf=17,capex=18,ervpsf=19,niy=20,exit=21,act1954=25,epc=27,comments=28)
def g(r,k): return ws.cell(r,C[k]).value
hdrmap={'A':'#','B':'Asset Name','C':'Region','D':'Sector','E':'Unit Number','F':'Tenant Name',
  'G':'Area GIA (sq ft)','H':'Entry Yield (NIY)','I':'Exit Yield','J':'Lease Start','K':'Lease Expiry',
  'L':'Break Date','M':'Break Taken (1=Yes,0=No)','N':'Rent Review / MTM Date',
  'O':'Event @ Expiry (Y=Renew / X=Vacate)','P':'Vacant @ Entry (Y/N)','S':'Passing Rent (£ pa)',
  'T':'ERV (£ pa)','U':'ERV (£ psf)','Y':'Assumed Void (mths)','Z':'Assumed Rent Free (mths)',
  'AA':'Re-letting Capex (£ psf)','AY':'1954 Act (Y/N)','AZ':'EPC Rating','BB':'Rent Review (Y/N)',
  'BC':'Term Certain (mths)','BE':'Rental Guarantee (Y/N)','BF':'Guarantee/Deduction Rent (£ pa)','BG':'Guarantee Period (mths)'}
out=openpyxl.Workbook(); o=out.active; o.title='Cannon RR'
for cc,lab in hdrmap.items(): o.cell(1,cix(cc)).value=lab

flags=[]; rows=[]; n=0
for r in range(2, ws.max_row+1):
    unit=g(r,'unit'); tenant=g(r,'tenant')
    if unit in (None,'') or str(unit).strip().upper()=='TOTAL': continue
    n+=1; orow=1+n
    traw=str(tenant or ''); cm=str(g(r,'comments') or ''); low=cm.lower()
    vacant='vacant' in traw.lower()
    under_offer='under offer' in traw.lower() or 'u/o' in low
    brk_taken=str(g(r,'brk') or '').strip().upper()=='Y'
    vacating=str(g(r,'vacating') or '').strip().upper()=='Y'
    area=mny(g(r,'gia')); ervpsf=mny(g(r,'ervpsf')); mrent=mny(g(r,'mrent_pa'))
    relet = (brk_taken or vacating) and not vacant      # mid-hold vacate/break re-let
    event='X' if (vacant or brk_taken or vacating) else 'Y'
    # guarantee psf/period from comment
    gpsf=None; gper=None
    mg=re.search(r'guarantee[^£%]*£?\s*([\d.]+)\s*psf', cm, re.I)
    if mg: gpsf=float(mg.group(1))
    mp=re.search(r'(\d+)\s*month', cm, re.I)
    if mp: gper=int(mp.group(1))
    J=pdate(g(r,'start')); K=pdate(g(r,'expiry')); Lb=pdate(g(r,'brkdate')); Nrv=pdate(g(r,'review'))
    a54=str(g(r,'act1954') or '').strip().lower()
    act='Y' if a54=='inside' else ('N' if a54=='outside' else None)
    void=mny(g(r,'lvoid')); rf=mny(g(r,'rf')); capex=mny(g(r,'capex'))
    erv_pa = ervpsf*area if (ervpsf and area) else None
    if vacant:
        S=0.0; P='Y'; T=erv_pa; U=ervpsf
        BF=(gpsf*area) if (gpsf and area) else erv_pa  # deduction/guarantee rent £pa
        BG=gper or 12; BE='Y'
        J=K=Lb=Nrv=None; Yv=Zv=None; AAcx=None; Mb=0; O='X'
    elif under_offer:
        S=mrent; P='N'; T=erv_pa; U=ervpsf
        J=ENTRY; K=EXITDATE; Lb=None; Mb=0; O='Y'; Nrv=None
        Yv=Zv=None; AAcx=None; BF=None; BG=None; BE='N'
    else:
        S=mrent; P='N'; T=erv_pa; U=ervpsf; O=event; Mb=1 if brk_taken else 0
        Lb=(Lb if brk_taken else None); Nrv=(Nrv if (Nrv and not brk_taken) else None)
        Yv=(void if void is not None else 12) if relet else None
        Zv=(rf if rf is not None else 6) if relet else None
        AAcx=(capex if capex is not None else 25) if relet else None
        BF=None; BG=None; BE='N'
    vals={'A':n,'B':ASSET,'C':REGION,'D':'Industrial','E':unit,'F':traw,'G':area,
      'H':mny(g(r,'niy')),'I':mny(g(r,'exit')),'J':J,'K':K,'L':Lb,'M':Mb,'N':Nrv,'O':O,'P':P,
      'S':S,'T':T,'U':U,'Y':Yv,'Z':Zv,'AA':AAcx,'AY':act,'AZ':g(r,'epc'),
      'BB':'Y' if (Nrv and not vacant) else 'N','BC':60,'BE':BE,'BF':BF,'BG':BG}
    for cc,val in vals.items(): o.cell(orow,cix(cc)).value=val
    rows.append((unit,traw,vals)); 
    if vacant: flags.append(f"U{n} {unit}: VACANT@entry; cap on ERV £{(T or 0):,.0f} ({ervpsf}psf), deduct guarantee £{(BF or 0):,.0f} ({gpsf or ervpsf}psf/{BG}mo) off PP; re-let I8/I9.")
    if under_offer: flags.append(f"U{n} {unit}: UNDER OFFER; let @£{mrent:,.0f}pa from entry, 5yr term (yr3 break NOT taken).")
    if brk_taken: flags.append(f"U{n} {unit}: BREAK taken {Lb.date() if Lb else '?'} -> vacate+re-let (void {Yv}/RF {Zv}/capex {AAcx}).")
    if vacating and not brk_taken: flags.append(f"U{n} {unit}: VACATE@expiry -> re-let (void {Yv}/RF {Zv}/capex {AAcx}).")
out.save(OUT)
print(f"Normalised {n} Cannon units -> {OUT}")
nv=sum(1 for u,t,v in rows if v['P']=='Y'); ncx=sum(1 for u,t,v in rows if v['AA']); nbrk=sum(1 for u,t,v in rows if v['M']==1)
print(f"vacant@entry={nv} | capex units={ncx} | break-taken={nbrk}")
print(f"\nFLAGS ({len(flags)}):")
for f in flags: print("  -",f)
