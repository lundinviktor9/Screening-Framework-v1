#!/usr/bin/env python3
"""normalise_rr.py — raw broker rent roll -> normalised template-column rent roll (v13).

Maps any broker layout (matched by header keywords) into the canonical 'Newbury RR'-style
sheet the injection consumes (template-letter columns A..BG, one row per unit). Applies the
deterministic v13 classification (break/vacate->event X; vacant->P=Y; review->BB) and the
house defaults; FLAGS judgement calls. Output sheet name 'Rent Roll'.

Usage: normalise_rr.py --raw RAW.xlsx [--sheet "Rent Roll"] --out NORM.xlsx
       [--asset "..."]  (asset name override; else built from Asset+Location)
"""
import argparse, openpyxl, datetime
from openpyxl.utils import get_column_letter as gl, column_index_from_string as cix

def find_cols(ws):
    """map header keyword -> column index, scanning the header row (row with most labels)."""
    hdr_row=max(range(1,min(ws.max_row,8)+1), key=lambda r: sum(1 for c in range(1,ws.max_column+1) if isinstance(ws.cell(r,c).value,str)))
    H={}
    for c in range(1,ws.max_column+1):
        v=ws.cell(hdr_row,c).value
        if isinstance(v,str): H[v.strip().lower()]=c
    def col(*keys):
        for k in keys:
            for h,c in H.items():
                if k in h: return c
        return None
    return hdr_row, col

def num(x):
    return x if isinstance(x,(int,float)) else None

def money(x):
    if isinstance(x,(int,float)): return float(x)
    if isinstance(x,str):
        t=x.replace('\u00a3','').replace(',','').replace(' ','').strip()
        try: return float(t)
        except: return None
    return None

import datetime as _dt
def pdate(x):
    if isinstance(x,_dt.datetime): return x
    if isinstance(x,_dt.date): return _dt.datetime(x.year,x.month,x.day)
    if isinstance(x,str):
        t=x.strip()
        for fmt in ('%d/%m/%Y','%d/%m/%y','%Y-%m-%d'):
            try: return _dt.datetime.strptime(t,fmt)
            except: pass
    return None

def main():
    ap=argparse.ArgumentParser(); ap.add_argument('--raw',required=True); ap.add_argument('--sheet',default=None)
    ap.add_argument('--out',required=True); ap.add_argument('--asset',default=None)
    a=ap.parse_args()
    wb=openpyxl.load_workbook(a.raw,data_only=True)
    ws=wb[a.sheet] if a.sheet else wb[wb.sheetnames[0]]
    hdr,col=find_cols(ws)
    c=dict(asset=col('asset'),loc=col('location','region'),unit=col('unit'),tenant=col('tenant'),
           gia=col('gia','area'),niy=col('entry niy','entry yield'),exit=col('exit niy','exit yield'),
           start=col('lease start'),review=col('review'),brkdate=col('break date'),brk=col('break (y'),
           expiry=col('lease expiry'),vac_exp=col('vacate at expiry'),hr_pa=col('headline rent pa'),
           erv=col('erv'),capex=col('capex'),void=col('void'),rf=col('rent fee'),lt=col('l&t'),
           epc=col('epc'),comments=col('comments'))
    out=openpyxl.Workbook(); o=out.active; o.title='Rent Roll'
    # header (normalised template-letter columns)
    hdrmap={'A':'#','B':'Asset Name','C':'Region','D':'Sector','E':'Unit Number','F':'Tenant Name',
            'G':'Area GIA (sq ft)','H':'Entry Yield (NIY)','I':'Exit Yield','J':'Lease Start','K':'Lease Expiry',
            'L':'Break Date','M':'Break Taken (1=Yes,0=No)','N':'Rent Review / MTM Date',
            'O':'Event @ Expiry (Y=Renew / X=Vacate)','P':'Vacant @ Entry (Y/N)','S':'Passing Rent (£ pa)',
            'T':'ERV (£ pa)','U':'ERV (£ psf)','Y':'Assumed Void (mths)','Z':'Assumed Rent Free (mths)',
            'AA':'Re-letting Capex (£ psf)','AY':'1954 Act (Y/N)','AZ':'EPC Rating','BB':'Rent Review (Y/N)',
            'BC':'Term Certain (mths)','BE':'Rental Guarantee (Y/N)','BF':'Guarantee/Headline Rent (£ pa)','BG':'Guarantee Period (mths)'}
    for cc,lab in hdrmap.items(): o.cell(1,cix(cc)).value=lab
    def g(r,key):
        idx=c[key]; return ws.cell(r,idx).value if idx else None
    flags=[]; n=0
    for r in range(hdr+1, ws.max_row+1):
        unit=g(r,'unit')
        if unit in (None,''): continue
        n+=1; orow=1+n
        tenant=(g(r,'tenant') or '')
        vacant = str(tenant).strip().lower()=='vacant'
        brk = str(g(r,'brk') or '').strip().upper()=='Y'
        vac_exp = g(r,'vac_exp') in (1,'1','Y','y',True)
        review = g(r,'review')
        ervpsf = money(g(r,'erv')); area=money(g(r,'gia'))
        hr_pa = money(g(r,'hr_pa'))
        event = 'X' if (vacant or brk or vac_exp) else 'Y'
        bb = 'Y' if (review is not None and not brk and not vacant) else 'N'
        asset = a.asset or (f"{g(r,'asset')}, {g(r,'loc')}" if g(r,'asset') and g(r,'loc') else g(r,'asset'))
        vals={'A':n,'B':asset,'C':g(r,'loc'),'D':'Industrial','E':unit,'F':tenant,'G':area,
              'H':num(g(r,'niy')),'I':num(g(r,'exit')),'J':pdate(g(r,'start')),'K':pdate(g(r,'expiry')),'L':(pdate(g(r,'brkdate')) if brk else None),
              'M':1 if brk else 0,'N':(pdate(review) if bb=='Y' else None),'O':event,'P':'Y' if vacant else 'N',
              'S':0 if vacant else hr_pa,'T':(ervpsf*area if ervpsf and area else None),'U':ervpsf,
              'Y':((num(g(r,'void')) if num(g(r,'void')) is not None else 9) if event=='X' else None),
              'Z':((num(g(r,'rf')) if num(g(r,'rf')) is not None else 6) if event=='X' else None),
              'AA':num(g(r,'capex')),'AY':g(r,'lt'),'AZ':g(r,'epc'),'BB':bb,'BC':60,
              'BE':'Y' if vacant else 'N','BF':hr_pa if vacant else None,'BG':12 if vacant else None}
        for cc,val in vals.items(): o.cell(orow,cix(cc)).value=val
        if vacant: flags.append(f"Unit {unit}: VACANT@entry — priced on headline rent £{(hr_pa or 0):,.0f} (basis: confirm)")
        if brk: flags.append(f"Unit {unit}: BREAK exercised -> vacate+re-let (confirm)")
        if vac_exp and not brk: flags.append(f"Unit {unit}: VACATE at expiry (confirm vs renew)")
        _cm=str(g(r,'comments') or '').lower()
        if (any(k in _cm for k in ['taw','holding over','tenant at will','renewal negotiat']) and not vacant and event=='Y'):
            flags.append(f"Unit {unit}: HOLDING-OVER/TAW — defaulted to RENEW@ERV; confirm vs vacate+re-let")
    out.save(a.out)
    print(f"normalised {n} units -> {a.out}")
    print("FLAGS for human confirmation:")
    for f in flags: print("  -", f)

if __name__=='__main__': main()
