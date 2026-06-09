"""
hand_adapters.py - known-broker fast path: importable, faithful ports of the proven
per-broker normalisers (Cannon, Meadow), plus a dispatcher used by normalise_auto.

Why this exists: the generic LLM/column-map path in normalise_auto maps COLUMNS only and
deliberately flags broker-specific judgement encodings (Cannon's Y/N break flags,
guarantee-rent-from-comments, under-offer dating; Meadow's multi-estate layout) rather than
auto-resolving them. For brokers whose layout we already know, the hand adapter reproduces the
trusted canonical RR exactly. These are line-for-line ports of normalise_cannon.py /
normalise_meadow.py wrapped as functions (the CLI scripts remain for standalone use).

Eval (2026-06-09): cannon 28/28, meadow 39/39 deterministic columns vs the trusted RRs; both
tie out through Mode B (checks.pass, anchor ok, 0 error cells).
"""
from __future__ import annotations
import datetime as _dt
import re as _re
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.utils import column_index_from_string as cix


def _pdate(x):
    if isinstance(x, _dt.datetime):
        return x
    if isinstance(x, _dt.date):
        return _dt.datetime(x.year, x.month, x.day)
    if isinstance(x, str):
        for f in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d'):
            try:
                return _dt.datetime.strptime(x.strip(), f)
            except Exception:
                pass
    return None


def _mny(x):
    return float(x) if isinstance(x, (int, float)) else None


# ---------------------------------------------------------------- Cannon

CANNON_SHEET = "Cannon Ind."
CANNON_ASSET = "Cannon Industrial Estate, Milton Keynes"
CANNON_REGION = "Milton Keynes"
CANNON_ENTRY = _dt.datetime(2026, 6, 30)
CANNON_EXIT = _dt.datetime(2031, 6, 30)


def normalise_cannon(src: str, out: str, asset: Optional[str] = None, region: Optional[str] = None,
                     entry: Optional[_dt.datetime] = None, exitdate: Optional[_dt.datetime] = None,
                     sheet: str = CANNON_SHEET, rr_sheet: str = "DealRR", **_) -> Dict[str, Any]:
    asset = asset or CANNON_ASSET
    region = region or CANNON_REGION
    ENTRY = entry or CANNON_ENTRY
    EXITDATE = exitdate or CANNON_EXIT
    ws = openpyxl.load_workbook(src, data_only=True)[sheet]
    C = dict(unit=1, tenant=2, gia=3, start=4, brkdate=5, review=6, expiry=7, brk=8, vacating=9,
             mrent_pa=12, lvoid=16, rf=17, capex=18, ervpsf=19, niy=20, exit=21, act1954=25, epc=27, comments=28)

    def g(r, k):
        return ws.cell(r, C[k]).value

    hdrmap = {'A': '#', 'B': 'Asset Name', 'C': 'Region', 'D': 'Sector', 'E': 'Unit Number', 'F': 'Tenant Name',
              'G': 'Area GIA (sq ft)', 'H': 'Entry Yield (NIY)', 'I': 'Exit Yield', 'J': 'Lease Start', 'K': 'Lease Expiry',
              'L': 'Break Date', 'M': 'Break Taken (1=Yes,0=No)', 'N': 'Rent Review / MTM Date',
              'O': 'Event @ Expiry (Y=Renew / X=Vacate)', 'P': 'Vacant @ Entry (Y/N)', 'S': 'Passing Rent (£ pa)',
              'T': 'ERV (£ pa)', 'U': 'ERV (£ psf)', 'Y': 'Assumed Void (mths)', 'Z': 'Assumed Rent Free (mths)',
              'AA': 'Re-letting Capex (£ psf)', 'AY': '1954 Act (Y/N)', 'AZ': 'EPC Rating', 'BB': 'Rent Review (Y/N)',
              'BC': 'Term Certain (mths)', 'BE': 'Rental Guarantee (Y/N)', 'BF': 'Guarantee/Deduction Rent (£ pa)', 'BG': 'Guarantee Period (mths)'}
    out_wb = openpyxl.Workbook()
    o = out_wb.active
    o.title = rr_sheet
    for cc, lab in hdrmap.items():
        o.cell(1, cix(cc)).value = lab

    flags: List[str] = []
    n = 0
    for r in range(2, ws.max_row + 1):
        unit = g(r, 'unit')
        tenant = g(r, 'tenant')
        if unit in (None, '') or str(unit).strip().upper() == 'TOTAL':
            continue
        n += 1
        orow = 1 + n
        traw = str(tenant or '')
        cm = str(g(r, 'comments') or '')
        low = cm.lower()
        vacant = 'vacant' in traw.lower()
        under_offer = 'under offer' in traw.lower() or 'u/o' in low
        brk_taken = str(g(r, 'brk') or '').strip().upper() == 'Y'
        vacating = str(g(r, 'vacating') or '').strip().upper() == 'Y'
        area = _mny(g(r, 'gia'))
        ervpsf = _mny(g(r, 'ervpsf'))
        mrent = _mny(g(r, 'mrent_pa'))
        relet = (brk_taken or vacating) and not vacant
        event = 'X' if (vacant or brk_taken or vacating) else 'Y'
        gpsf = None
        gper = None
        mg = _re.search(r'guarantee[^£%]*£?\s*([\d.]+)\s*psf', cm, _re.I)
        if mg:
            gpsf = float(mg.group(1))
        mp = _re.search(r'(\d+)\s*month', cm, _re.I)
        if mp:
            gper = int(mp.group(1))
        J = _pdate(g(r, 'start'))
        K = _pdate(g(r, 'expiry'))
        Lb = _pdate(g(r, 'brkdate'))
        Nrv = _pdate(g(r, 'review'))
        a54 = str(g(r, 'act1954') or '').strip().lower()
        act = 'Y' if a54 == 'inside' else ('N' if a54 == 'outside' else None)
        void = _mny(g(r, 'lvoid'))
        rf = _mny(g(r, 'rf'))
        capex = _mny(g(r, 'capex'))
        erv_pa = ervpsf * area if (ervpsf and area) else None
        if vacant:
            S = 0.0; P = 'Y'; T = erv_pa; U = ervpsf
            BF = (gpsf * area) if (gpsf and area) else erv_pa
            BG = gper or 12; BE = 'Y'
            J = K = Lb = Nrv = None; Yv = Zv = None; AAcx = None; Mb = 0; O = 'X'
        elif under_offer:
            S = mrent; P = 'N'; T = erv_pa; U = ervpsf
            J = ENTRY; K = EXITDATE; Lb = None; Mb = 0; O = 'Y'; Nrv = None
            Yv = Zv = None; AAcx = None; BF = None; BG = None; BE = 'N'
        else:
            S = mrent; P = 'N'; T = erv_pa; U = ervpsf; O = event; Mb = 1 if brk_taken else 0
            Lb = (Lb if brk_taken else None); Nrv = (Nrv if (Nrv and not brk_taken) else None)
            Yv = (void if void is not None else 12) if relet else None
            Zv = (rf if rf is not None else 6) if relet else None
            AAcx = (capex if capex is not None else 25) if relet else None
            BF = None; BG = None; BE = 'N'
        vals = {'A': n, 'B': asset, 'C': region, 'D': 'Industrial', 'E': unit, 'F': traw, 'G': area,
                'H': _mny(g(r, 'niy')), 'I': _mny(g(r, 'exit')), 'J': J, 'K': K, 'L': Lb, 'M': Mb, 'N': Nrv, 'O': O, 'P': P,
                'S': S, 'T': T, 'U': U, 'Y': Yv, 'Z': Zv, 'AA': AAcx, 'AY': act, 'AZ': g(r, 'epc'),
                'BB': 'Y' if (Nrv and not vacant) else 'N', 'BC': 60, 'BE': BE, 'BF': BF, 'BG': BG}
        for cc, val in vals.items():
            o.cell(orow, cix(cc)).value = val
        if vacant:
            flags.append({"unit": unit, "signal": "vacant@entry", "needs_signoff": True,
                          "note": f"cap on ERV; deduct guarantee off PP; re-let I8/I9"})
        if under_offer:
            flags.append({"unit": unit, "signal": "under_offer", "needs_signoff": True,
                          "note": "let from entry, 5yr term"})
        if brk_taken:
            flags.append({"unit": unit, "signal": "break_taken", "needs_signoff": True,
                          "note": "vacate + re-let"})
        if vacating and not brk_taken:
            flags.append({"unit": unit, "signal": "vacate@expiry", "needs_signoff": True,
                          "note": "re-let"})
    out_wb.save(out)
    return {"rr_xlsx": out, "rr_sheet": rr_sheet, "units": n, "flags": flags}


# ---------------------------------------------------------------- Meadow

MEADOW_SHEET = "Meadow "
MEADOW_REGIONS = {
    "Zodiac Park, West Drayton": "South East",
    "Basildon Trade Centre": "South East",
    "Hobley Drive, Swindon": "South West",
}


def normalise_meadow(src: str, out: str, asset: Optional[str] = None, region: str = "UK",
                     asset_filter: Optional[str] = None, region_by_asset: Optional[Dict[str, str]] = None,
                     sheet: str = MEADOW_SHEET, rr_sheet: str = "DealRR", **_) -> Dict[str, Any]:
    """All estates by default (asset_filter=None). Per-estate region from region_by_asset
    (defaults to the known Meadow map), else the `region` fallback."""
    region_by_asset = region_by_asset or MEADOW_REGIONS
    ws = openpyxl.load_workbook(src, data_only=True)[sheet]
    M = dict(asset='A', region='B', unit='C', tenant='D', gia='E', eyld='G', xyld='H', start='L', expiry='M',
             brkdate='O', brktaken='P', review='Q', event='AJ', passing='W', ervpa='AA', ervpsf='AC',
             rv='AD', rates='AE', sc='AF', void='AU', rf='AV', capex='AR')

    def g(r, k):
        return ws.cell(r, cix(M[k])).value

    hdrmap = {'A': '#', 'B': 'Asset Name', 'C': 'Region', 'D': 'Sector', 'E': 'Unit Number', 'F': 'Tenant Name',
              'G': 'Area GIA (sq ft)', 'H': 'Entry Yield (NIY)', 'I': 'Exit Yield', 'J': 'Lease Start', 'K': 'Lease Expiry',
              'L': 'Break Date', 'M': 'Break Taken (1=Yes,0=No)', 'N': 'Rent Review / MTM Date',
              'O': 'Event @ Expiry (Y=Renew / X=Vacate)', 'P': 'Vacant @ Entry (Y/N)', 'S': 'Passing Rent (£ pa)',
              'T': 'ERV (£ pa)', 'U': 'ERV (£ psf)', 'V': 'Rateable Value', 'W': 'Business Rates (£ pa)', 'X': 'Service Charge (£ pa)',
              'Y': 'Assumed Void (mths)', 'Z': 'Assumed Rent Free (mths)', 'AA': 'Re-letting Capex (£ psf)',
              'BB': 'Rent Review (Y/N)', 'BC': 'Term Certain (mths)', 'BE': 'Rental Guarantee (Y/N)', 'BF': 'Guarantee/Deduction Rent (£ pa)', 'BG': 'Guarantee Period (mths)'}
    out_wb = openpyxl.Workbook()
    o = out_wb.active
    o.title = rr_sheet
    for c, l in hdrmap.items():
        o.cell(1, cix(c)).value = l

    n = 0
    flags: List[str] = []
    for r in range(2, ws.max_row + 1):
        a = g(r, 'asset')
        if not a or 'Total' in str(a):
            continue
        if asset_filter and asset_filter.lower() not in str(a).lower():
            continue
        reg = region_by_asset.get(str(a), region)
        n += 1
        orow = 1 + n
        tenant = str(g(r, 'tenant') or '')
        vacant = 'vacant' in tenant.lower()
        brk = (g(r, 'brktaken') in (1, '1')) and not ('vacant' in tenant.lower())
        event = str(g(r, 'event') or 'Y').strip().upper()
        if vacant:
            event = 'X'
        area = _mny(g(r, 'gia'))
        ervpsf = _mny(g(r, 'ervpsf'))
        ervpa = _mny(g(r, 'ervpa'))
        relet = (event == 'X') and not vacant
        Nrev = _pdate(g(r, 'review'))
        vals = {'A': n, 'B': a, 'C': reg, 'D': 'Industrial', 'E': g(r, 'unit'), 'F': tenant, 'G': area,
                'H': _mny(g(r, 'eyld')), 'I': _mny(g(r, 'xyld')),
                'J': None if vacant else _pdate(g(r, 'start')), 'K': None if vacant else _pdate(g(r, 'expiry')),
                'L': _pdate(g(r, 'brkdate')) if brk else None, 'M': 1 if brk else 0,
                'N': Nrev if (Nrev and not vacant) else None, 'O': event, 'P': 'Y' if vacant else 'N',
                'S': 0 if vacant else _mny(g(r, 'passing')), 'T': ervpa, 'U': ervpsf,
                'V': _mny(g(r, 'rv')), 'W': _mny(g(r, 'rates')), 'X': _mny(g(r, 'sc')),
                'Y': ((_mny(g(r, 'void')) if _mny(g(r, 'void')) is not None else 9) if relet else None),
                'Z': ((_mny(g(r, 'rf')) if _mny(g(r, 'rf')) is not None else 6) if relet else None),
                'AA': ((_mny(g(r, 'capex')) if _mny(g(r, 'capex')) is not None else 20) if relet else None),
                'BB': 'Y' if (Nrev and not vacant) else 'N', 'BC': 60,
                'BE': 'Y' if vacant else 'N', 'BF': ervpa if vacant else None, 'BG': 12 if vacant else None}
        for c, v in vals.items():
            o.cell(orow, cix(c)).value = v
        if vacant:
            flags.append({"unit": vals['E'], "signal": "vacant@entry", "needs_signoff": True, "note": "ERV cap, re-let I8/I9"})
        if brk:
            flags.append({"unit": vals['E'], "signal": "break_taken", "needs_signoff": True, "note": "vacate + re-let"})
        if event == 'X' and not brk and not vacant:
            flags.append({"unit": vals['E'], "signal": "vacate@expiry", "needs_signoff": True, "note": "re-let"})
    out_wb.save(out)
    return {"rr_xlsx": out, "rr_sheet": rr_sheet, "units": n, "flags": flags}


# ---------------------------------------------------------------- dispatch

# raw-sheet signature -> adapter. First match wins.
HAND_ADAPTERS = [
    {"name": "cannon", "sheet": CANNON_SHEET, "fn": normalise_cannon},
    {"name": "meadow", "sheet": MEADOW_SHEET, "fn": normalise_meadow},
]

# A full MLI model embeds every broker's raw sheet, so plain sheet-name detection would
# mis-fire on one. Real uploads are STANDALONE rent rolls; skip dispatch if the file looks
# like a populated model.
_MODEL_MARKERS = {"Cash Flow Output", "Global Assumptions", "TENANCY SCHEDULE (template)"}


def dispatch_hand_adapter(file_path: str, asset: Optional[str], region: Optional[str],
                          out_xlsx: str, rr_sheet: str = "DealRR") -> Optional[Dict[str, Any]]:
    """If the uploaded workbook carries a known broker's raw sheet, run that hand adapter and
    return a normalise_auto-shaped result; else None (caller falls back to known-map / LLM)."""
    names = openpyxl.load_workbook(file_path, read_only=True).sheetnames
    if _MODEL_MARKERS & set(names):
        return None  # full model, not a standalone rent roll
    for h in HAND_ADAPTERS:
        if h["sheet"] in names:
            res = h["fn"](file_path, out_xlsx, asset=asset, region=region or "UK", rr_sheet=rr_sheet)
            return {
                "rr_xlsx": res["rr_xlsx"], "rr_sheet": res["rr_sheet"], "units": res["units"],
                "flags": res["flags"],
                "mapping": {"_source": f"hand:{h['name']}", "sheet": h["sheet"]},
            }
    return None
