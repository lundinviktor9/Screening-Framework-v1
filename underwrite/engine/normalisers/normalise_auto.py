"""
normalise_auto.py - layout-agnostic rent-roll normaliser (hybrid LLM-map + deterministic-extract).

ANY broker tenancy schedule -> the canonical field-dictionary layout that the underwrite engine
consumes. Two stages, deliberately separated so financial values are NEVER produced by an LLM:

  Stage 1 (LLM, fuzzy):  propose_mapping(sheet_preview) -> a MAPPING (which sheet, header row,
                         data range, {canonical field -> source column}, derivations, flags).
                         The LLM reads HEADERS + a few sample rows and reasons about layout only.
  Stage 2 (code, exact): apply_mapping(ws, mapping) reads the ACTUAL cells via the mapping and
                         parses them deterministically (dates, money, %, psf<->pa), applies the
                         house derivations, validates vs tenancy_schedule.schema.json, and writes
                         the canonical RR sheet. Numbers come straight from cells.

Judgment fields (entry/exit yield = pricing; area reconciliation; vacate-at-expiry house view)
are NOT auto-resolved here - they are emitted as flags for analyst sign-off (Mode A).
"""
from __future__ import annotations
import datetime as _dt
import re
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.utils import column_index_from_string as cix, get_column_letter as gl

# ----- deterministic parsers (the trustworthy half) -----

def parse_num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^0-9.\-]", "", str(v).replace(",", ""))
    return float(s) if s not in ("", "-", ".") else None

def parse_pct(v: Any) -> Optional[float]:
    """'7.0%'/7.0 -> 0.07 ; 0.07 -> 0.07."""
    n = parse_num(v)
    if n is None:
        return None
    if isinstance(v, str) and "%" in v:
        return n / 100.0
    return n / 100.0 if n > 1.0 else n   # 7.25 -> 0.0725 ; 0.0725 -> 0.0725

def parse_date(v: Any) -> Optional[_dt.datetime]:
    if v is None or v == "":
        return None
    if isinstance(v, _dt.datetime):
        return _dt.datetime(v.year, v.month, v.day)
    if isinstance(v, _dt.date):
        return _dt.datetime(v.year, v.month, v.day)
    try:
        from dateutil import parser as _p          # type: ignore
        return _p.parse(str(v), dayfirst=True).replace(hour=0, minute=0, second=0, microsecond=0)
    except Exception:
        return None

def parse_yn(v: Any) -> str:
    s = str(v or "").strip().lower()
    return "Y" if s in ("y", "yes", "1", "true", "vacant") else "N"

# ----- mapping-driven extraction (Stage 2) -----

# canonical fields the engine RR needs (field-dictionary letters in the RR sheet)
RR_COLS = {  # canonical name -> RR output column letter
    "Asset Name": "B", "Region": "C", "Sector": "D", "Unit Number": "E", "Tenant Name": "F",
    "Area GIA (sq ft)": "G", "Entry Yield (NIY)": "H", "Exit Yield": "I", "Lease Start": "J",
    "Lease Expiry": "K", "Break Date": "L", "Break Taken (1=Yes,0=No)": "M",
    "Rent Review / MTM Date": "N", "Event @ Expiry (Y=Renew / X=Vacate)": "O",
    "Vacant @ Entry (Y/N)": "P", "Passing Rent (pa)": "S", "ERV (pa)": "T", "ERV (psf)": "U",
    "Assumed Void (mths)": "Y", "Assumed Rent Free (mths)": "Z", "Re-letting Capex (psf)": "AA",
    "Guarantee Rent (pa)": "BF",
}
NUM = {"Area GIA (sq ft)", "Passing Rent (pa)", "ERV (pa)", "ERV (psf)", "Assumed Void (mths)",
       "Assumed Rent Free (mths)", "Re-letting Capex (psf)", "Guarantee Rent (pa)"}
PCT = {"Entry Yield (NIY)", "Exit Yield"}
DATE = {"Lease Start", "Lease Expiry", "Break Date", "Rent Review / MTM Date"}
YN = {"Vacant @ Entry (Y/N)"}


def apply_mapping(ws, mapping: Dict[str, Any], asset: str, region: str) -> Dict[str, Any]:
    """Read cells via the mapping, parse + derive deterministically, return canonical rows + flags."""
    colmap: Dict[str, str] = mapping["columns"]            # canonical field -> source col letter
    aux: Dict[str, str] = mapping.get("aux_columns", {})   # extra source cols used by derivations
    hdr = int(mapping["data_start_row"])
    unit_src = colmap["Unit Number"]
    rows: List[Dict[str, Any]] = []
    flags: List[Dict[str, Any]] = []

    def cell(r, field_or_letter):
        col = colmap.get(field_or_letter) or aux.get(field_or_letter)
        return ws.cell(r, cix(col)).value if col else None

    r = hdr
    while ws.cell(r, cix(unit_src)).value not in (None, ""):
        row: Dict[str, Any] = {"Asset Name": asset, "Region": region, "Sector": "Industrial"}
        for field, col in colmap.items():
            raw = ws.cell(r, cix(col)).value
            if field in PCT:
                row[field] = parse_pct(raw)
            elif field in NUM:
                row[field] = parse_num(raw)
            elif field in DATE:
                row[field] = parse_date(raw)
            elif field in YN:
                row[field] = parse_yn(raw)
            else:
                row[field] = raw
        # ----- house derivations (deterministic) -----
        # ERV pa from psf * area when pa absent
        if not row.get("ERV (pa)") and row.get("ERV (psf)") and row.get("Area GIA (sq ft)"):
            row["ERV (pa)"] = round(row["ERV (psf)"] * row["Area GIA (sq ft)"], 2)
        # vacant @ entry: explicit flag, else tenant 'Vacant' / passing 0
        if "Vacant @ Entry (Y/N)" not in colmap:
            t = str(row.get("Tenant Name") or "").lower()
            row["Vacant @ Entry (Y/N)"] = "Y" if (t == "vacant" or not row.get("Passing Rent (pa)")) else "N"
        # event @ expiry: X (vacate) if break taken OR vacate-at-expiry flag, else Y (renew)
        bt = str(cell(r, "Break Taken (1=Yes,0=No)") or "").strip() in ("1", "1.0")
        vac_exp = str(cell(r, "_vacate_at_expiry") or "").strip() in ("1", "1.0", "y", "yes")
        row["Event @ Expiry (Y=Renew / X=Vacate)"] = "X" if (bt or vac_exp) else "Y"
        # guarantee rent: only meaningful for vacant@entry units
        if row.get("Vacant @ Entry (Y/N)") != "Y":
            row["Guarantee Rent (pa)"] = None
        rows.append(row)
        r += 1

    # ----- judgment flags (NOT auto-resolved) -----
    flags.append({"field": "Entry Yield (NIY)", "note": "schedule NIY is a starting point; confirm "
                  "the acquisition yield (pricing decision)", "needs_signoff": True})
    flags.append({"field": "Exit Yield", "note": "confirm exit yield (biggest value driver)",
                  "needs_signoff": True})
    for row in rows:
        if row.get("Vacant @ Entry (Y/N)") == "Y":
            flags.append({"unit": row["Unit Number"], "field": "Vacant @ Entry",
                          "note": "confirm vacant valuation basis (capitalise guarantee/headline)",
                          "needs_signoff": True})
    return {"rows": rows, "flags": flags, "units": len(rows)}


def write_rr_sheet(rows: List[Dict[str, Any]], out_xlsx: str, rr_sheet: str = "DealRR") -> None:
    """Write canonical rows to an xlsx RR sheet in the field-dictionary layout (row 1 = headers)."""
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = rr_sheet
    ws["A1"] = "#"
    for field, col in RR_COLS.items():
        ws[f"{col}1"] = field
    for i, row in enumerate(rows):
        rr = 2 + i
        ws.cell(rr, 1, i + 1)
        for field, col in RR_COLS.items():
            v = row.get(field)
            if v is not None:
                ws.cell(rr, cix(col), v)
    wb.save(out_xlsx)


# ----- Stage 1: LLM column mapping (run in the app, with the Anthropic API) -----

def sheet_preview(ws, n_rows: int = 20, n_cols: int = 40) -> str:
    """A compact grid the LLM can reason over: column letters + the first n_rows non-empty rows.

    Scans 20 rows by default (not 8): real broker schedules push the header down past
    title/logo/date/notes rows - e.g. Newbury's header sits on row 11 - so a short window
    misses the labels entirely. Empty rows are skipped but row numbers are preserved so the
    model can pin header_row / data_start_row precisely."""
    lines = []
    for r in range(1, n_rows + 1):
        cells = [f"{gl(c)}={ws.cell(r, c).value!r}" for c in range(1, n_cols + 1)
                 if ws.cell(r, c).value not in (None, "")]
        if cells:
            lines.append(f"row{r}: " + " | ".join(cells))
    return "\n".join(lines)

# ----- known-broker fast path (deterministic; skips the LLM when a layout is recognised) -----

def _normkey(x: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(x).lower())

def find_header_row(ws, tokens, min_hits: int = 4, max_scan: int = 30) -> Optional[int]:
    """First row (within max_scan) whose cells match >= min_hits of the signature tokens.
    Matching is on alphanumerics-only keys so 'GIA\\n(sq ft)' matches 'GIA (sq ft)'."""
    toks = {_normkey(t) for t in tokens}
    for r in range(1, min(ws.max_row, max_scan) + 1):
        cells = {_normkey(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)
                 if ws.cell(r, c).value not in (None, "")}
        if len(toks & cells) >= min_hits:
            return r
    return None

# Each entry: a signature (header labels to detect) + a fixed canonical->source-column map.
# Header row is FOUND (not hardcoded) so minor top-matter shifts don't break the match.
KNOWN_LAYOUTS = [
    {
        "name": "newbury_ts_new",
        "signature": ["Unit", "GIA (sq ft)", "Vacant @ Entry (Y/N)", "Passing Rent PA",
                      "Entry Yield", "Exit Yield"],
        "columns": {
            "Unit Number": "D", "Tenant Name": "E", "Area GIA (sq ft)": "F",
            "Vacant @ Entry (Y/N)": "G", "Lease Start": "H", "Rent Review / MTM Date": "I",
            "Break Date": "J", "Break Taken (1=Yes,0=No)": "K", "Lease Expiry": "L",
            "Passing Rent (pa)": "O", "Guarantee Rent (pa)": "Q", "Entry Yield (NIY)": "S",
            "Exit Yield": "T", "ERV (psf)": "V", "Re-letting Capex (psf)": "W",
            "Assumed Void (mths)": "X", "Assumed Rent Free (mths)": "Y",
        },
        "aux_columns": {"_vacate_at_expiry": "U"},
    },
]

def detect_known_mapping(ws) -> Optional[Dict[str, Any]]:
    """Return a ready mapping if the sheet matches a known broker layout, else None."""
    for lay in KNOWN_LAYOUTS:
        min_hits = max(4, len(lay["signature"]) // 2)
        hr = find_header_row(ws, lay["signature"], min_hits=min_hits)
        if hr:
            return {
                "sheet": ws.title, "header_row": hr, "data_start_row": hr + 1,
                "columns": dict(lay["columns"]), "aux_columns": dict(lay.get("aux_columns", {})),
                "flags": [{"field": "_source",
                           "note": f"known layout '{lay['name']}' matched (header row {hr})"}],
                "_source": f"known:{lay['name']}",
            }
    return None


MAPPING_PROMPT = """You map a broker rent roll to a canonical schema. Output STRICT JSON only.

Canonical fields (map each to the source COLUMN LETTER that holds it; omit if not present):
  Unit Number, Tenant Name, Area GIA (sq ft), Entry Yield (NIY), Exit Yield, Lease Start,
  Lease Expiry, Break Date, Break Taken (1=Yes,0=No), Rent Review / MTM Date,
  Vacant @ Entry (Y/N), Passing Rent (pa), Guarantee Rent (pa) [= headline rent for vacant units],
  ERV (pa), ERV (psf), Assumed Void (mths), Assumed Rent Free (mths), Re-letting Capex (psf)

Also return aux_columns for any "Vacate at Expiry" flag as key "_vacate_at_expiry".

CRITICAL RULES:
1. The header row is usually NOT row 1. Broker schedules begin with a title / address / "Today's
   Date" / map-link / notes block; the real header is the first row whose cells are column LABELS
   like "Unit", "GIA (sq ft)", "Passing Rent", "Entry Yield". Set "header_row" to that row and
   "data_start_row" to the first row of actual unit data beneath it.
2. Headers may span multiple rows or contain line breaks (e.g. "Tenant\\n[Guarantor]",
   "GIA\\n(sq ft)"). Read the combined label; map by meaning.
3. Rents may be quoted per-square-foot only. If you see a "psf" rent/ERV column but no annual (pa)
   column, map it to "ERV (psf)" (or note Passing is psf-only) - the code derives pa = psf x area.
4. Dates may be text ("Mar-25", "Q2 2026") rather than real dates - still map the column; parsing
   is done downstream, deterministically.
5. Ignore total / subtotal / footnote / blank-unit rows; data ends at the first such row.
6. Do NOT read or transcribe any numeric/date VALUES - only identify which column LETTER holds each
   field. Flag anything ambiguous (e.g. two plausible rent columns, missing yields, merged cells).

Schedule preview (column letters + first non-empty rows):
{preview}

Return JSON: {{"sheet": "<name>", "header_row": <int>, "data_start_row": <int>,
"columns": {{"<canonical field>": "<col letter>", ...}}, "aux_columns": {{...}},
"flags": [{{"field": "...", "note": "..."}}]}}"""

def propose_mapping(ws, sheet_name: str, anthropic_client=None, model: str = "claude-sonnet-4-6") -> Dict[str, Any]:
    """Ask the LLM for the column mapping. Requires an Anthropic client (app supplies it)."""
    import json
    prompt = MAPPING_PROMPT.format(preview=sheet_preview(ws))
    if anthropic_client is None:
        import anthropic  # type: ignore
        anthropic_client = anthropic.Anthropic()
    msg = anthropic_client.messages.create(
        model=model, max_tokens=1500,
        messages=[{"role": "user", "content": prompt}],
    )
    text = msg.content[0].text
    text = text[text.find("{"): text.rfind("}") + 1]
    m = json.loads(text)
    m.setdefault("sheet", sheet_name)
    return m


def normalise_auto(file_path: str, asset: str, region: str, out_xlsx: str,
                   rr_sheet: str = "DealRR", mapping: Optional[Dict[str, Any]] = None,
                   anthropic_client=None, sheet: Optional[str] = None) -> Dict[str, Any]:
    """End-to-end: resolve a mapping -> (code) extract + derive + write canonical RR.

    Mapping resolution order when `mapping` is not supplied:
      1. known-broker fast path (detect_known_mapping) - deterministic, no LLM;
      2. LLM propose_mapping for unknown layouts.
    Pass `mapping` explicitly to skip resolution (e.g. an analyst-confirmed map). `sheet` picks
    the source worksheet (defaults to the first sheet)."""
    if mapping is None:
        # known-broker hand adapter first (reproduces broker-specific judgement encodings the
        # generic mapper would only flag); falls through to known-map / LLM for unknown layouts.
        from .hand_adapters import dispatch_hand_adapter
        hand = dispatch_hand_adapter(file_path, asset, region, out_xlsx, rr_sheet)
        if hand is not None:
            return hand
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws0 = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    if mapping is None:
        mapping = detect_known_mapping(ws0)
        if mapping is None:
            mapping = propose_mapping(ws0, ws0.title, anthropic_client)
    ws = wb[mapping.get("sheet", ws0.title)]
    res = apply_mapping(ws, mapping, asset, region)
    write_rr_sheet(res["rows"], out_xlsx, rr_sheet)
    return {"rr_xlsx": out_xlsx, "rr_sheet": rr_sheet, "units": res["units"],
            "flags": res["flags"], "mapping": mapping}
