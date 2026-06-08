#!/usr/bin/env python3
"""
validate_and_write.py - write normalised tenancy rows into the
`TENANCY SCHEDULE (template)` sheet, with validation + an Excel sign-off checklist.

INPUT: a CSV whose header row is TEMPLATE COLUMN LETTERS (A,B,C,...,BG) - one data
row per unit. Claude produces this CSV by mapping the broker schedule against
references/field_dictionary.md. Empty cells are fine; leave a column out entirely
if the broker schedule has no data for it. Write dates as ISO YYYY-MM-DD.

USAGE:
  python validate_and_write.py --input normalised.csv \
      --model "Brunswick MLI portfolio - REBUILD v12.xlsx" \
      --out   "Brunswick MLI portfolio - <DEAL>.xlsx"
  # or standalone (no model): python validate_and_write.py --input normalised.csv --out filled.xlsx

Always writes a markdown report next to --out and prints it. NEVER recalc the output
through LibreOffice to "fix" dates - open it in Excel (see field_dictionary).
"""
import argparse, csv, os, sys, shutil, re
import datetime as dt
try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl --break-system-packages")

TPL_SHEET = "TENANCY SCHEDULE (template)"
FIRST_DATA_ROW = 24
REGISTER_RANGE = ("Cash Flow Output", "P", 5, 30)
DATE_COLS = ["J","K","L","N"] + ["A"+c for c in "LMNOPQRSTU"]
NUM_COLS = ["G","H","I","Q","R","S","T","U","V","W","X","Y","Z","AA","BA","BC","BD","BF","BG"] + ["A"+c for c in "BCDEFGHIJK"]
YN_COLS = ["M","O","P","AY","BB","BE"]
REQUIRED = ["B","E","G","H","I","J","K","N","O","P","S","T"]

def col_to_idx(letter):
    n=0
    for ch in letter: n=n*26+(ord(ch)-64)
    return n

def parse_date(v):
    """ISO YYYY-MM-DD parsed year-first (unambiguous); everything else day-first (UK)."""
    if v in (None,"") or (isinstance(v,float) and v!=v): return None
    if isinstance(v,(dt.datetime,dt.date)): return v
    s=str(v).strip()
    if not s: return None
    from dateutil import parser as dparser
    iso=bool(re.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}", s))
    try:
        return dparser.parse(s, dayfirst=not iso, yearfirst=iso)
    except Exception:
        for fmt in ("%Y-%m-%d","%d/%m/%Y","%d-%b-%y","%d-%b-%Y","%b-%y","%d.%m.%Y"):
            try: return dt.datetime.strptime(s,fmt)
            except Exception: pass
    return "UNPARSEABLE:"+s

def parse_num(v):
    if v in (None,""): return None
    if isinstance(v,(int,float)): return v
    raw=str(v).strip(); s=raw.replace(",","").replace("GBP","").replace("£","").replace("%","").strip()
    if s=="": return None
    try:
        x=float(s)
        if "%" in raw: x=x/100.0
        return x
    except Exception:
        return "NONNUMERIC:"+raw

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--input",required=True)
    ap.add_argument("--model",default=None)
    ap.add_argument("--out",required=True)
    ap.add_argument("--start-row",type=int,default=FIRST_DATA_ROW)
    a=ap.parse_args()
    rows=[]
    with open(a.input,newline="",encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            rows.append({(k or "").strip().upper():v for k,v in r.items()})
    report=["# Normalisation report - "+os.path.basename(a.out),"",
            "Units: **%d** | source: `%s`"%(len(rows),a.input),""]
    errors,warns,notes=[],[],[]
    register=set()
    if a.model:
        wbm=openpyxl.load_workbook(a.model,data_only=True)
        sh,colL,r0,r1=REGISTER_RANGE
        if sh in wbm.sheetnames:
            ws=wbm[sh]; ci=col_to_idx(colL)
            for rr in range(r0,r1+1):
                nm=ws.cell(row=rr,column=ci).value; tg=ws.cell(row=rr,column=ci+1).value
                if isinstance(nm,str) and nm.strip() and tg in (0,1,0.0,1.0): register.add(nm.strip())
    if a.model:
        shutil.copy(a.model,a.out); wb=openpyxl.load_workbook(a.out)
    else:
        wb=openpyxl.Workbook(); wb.active.title=TPL_SHEET
    if TPL_SHEET not in wb.sheetnames:
        errors.append("target has no '%s' sheet"%TPL_SHEET); _finish(report,errors,warns,notes,a); return
    tpl=wb[TPL_SHEET]
    if a.model:
        for rr in range(a.start_row,tpl.max_row+1):
            for cc in range(1,col_to_idx("BG")+1): tpl.cell(row=rr,column=cc).value=None
    assets=set()
    for i,row in enumerate(rows):
        out_r=a.start_row+i; uid="unit %d"%(i+1)
        for col,val in row.items():
            if val in (None,""): continue
            if col in DATE_COLS:
                p=parse_date(val)
                if isinstance(p,str) and p.startswith("UNPARSEABLE"):
                    errors.append("%s col %s: cannot parse date '%s'"%(uid,col,val)); continue
                tpl.cell(row=out_r,column=col_to_idx(col)).value=p
            elif col in NUM_COLS:
                p=parse_num(val)
                if isinstance(p,str) and p.startswith("NONNUMERIC"):
                    errors.append("%s col %s: non-numeric '%s'"%(uid,col,val)); continue
                tpl.cell(row=out_r,column=col_to_idx(col)).value=p
            else:
                tpl.cell(row=out_r,column=col_to_idx(col)).value=str(val).strip()
        g=parse_num(row.get("G")); s=parse_num(row.get("S")); t=parse_num(row.get("T"))
        h=parse_num(row.get("H")); ii=parse_num(row.get("I"))
        for req in REQUIRED:
            if row.get(req) in (None,""): warns.append("%s: missing REQUIRED col %s"%(uid,req))
        if isinstance(g,(int,float)):
            if g<=0: errors.append("%s: area G<=0"%uid)
            elif g<100 or g>1_000_000: warns.append("%s: area %0.0f sqft implausible"%(uid,g))
            if isinstance(s,(int,float)) and s>0:
                psf=s/g
                if psf<1 or psf>50: warns.append("%s: passing GBP%0.1f/sqft outside 1-50 band"%(uid,psf))
        if isinstance(s,(int,float)) and isinstance(t,(int,float)) and t>0 and s>0 and t<0.8*s:
            warns.append("%s: ERV %0.0f well below passing %0.0f (check)"%(uid,t,s))
        for y,nm in ((h,"entry"),(ii,"exit")):
            if isinstance(y,(int,float)) and not(0.03<=y<=0.12): warns.append("%s: %s yield %0.3f outside 3-12pct"%(uid,nm,y))
        js,ks,ls,ns=[parse_date(row.get(c)) for c in ("J","K","L","N")]
        isd=lambda x: isinstance(x,(dt.datetime,dt.date))
        if isd(js) and isd(ks) and ks<=js: errors.append("%s: expiry K not after start J"%uid)
        if isd(js) and isd(ls) and isd(ks) and not(js<ls<ks): warns.append("%s: break L not between start & expiry"%uid)
        if isd(js) and isd(ns) and isd(ks) and not(js<=ns<=ks): warns.append("%s: review N outside lease term"%uid)
        p=str(row.get("P","")).strip().upper()
        if p=="Y" and isinstance(s,(int,float)) and s>0: warns.append("%s: Vacant@Entry=Y but passing>0"%uid)
        if p=="N" and isinstance(s,(int,float)) and s==0: warns.append("%s: Vacant@Entry=N but passing=0"%uid)
        if row.get("B"): assets.add(str(row["B"]).strip())
    if a.model:
        missing=[x for x in assets if x not in register]
        if missing:
            errors.append("ASSET NAMES NOT IN REGISTER (units EXCLUDED until added to Cash Flow Output!P5:P30 with a 1/0 toggle): "+", ".join(sorted(missing)))
        notes.append("Distinct assets in input: "+", ".join(sorted(assets)))
        notes.append("Register currently holds: "+(", ".join(sorted(register)) or "(none read)"))
    notes.append("Wrote %d unit rows into '%s' rows %d-%d."%(len(rows),TPL_SHEET,a.start_row,a.start_row+len(rows)-1))
    if a.model and len(rows)>152:
        warns.append("%d units > 152: TI per-unit extract block (rows 35:186) + Include col may need extending - hand to runner."%len(rows))
    wb.save(a.out)
    _finish(report,errors,warns,notes,a)

def _finish(report,errors,warns,notes,a):
    report+=["## Validation","","- Errors: **%d**   Warnings: **%d**"%(len(errors),len(warns)),""]
    if errors: report+=["### ERRORS (fix before underwriting)"]+["- [ERROR] "+e for e in errors]+[""]
    if warns:  report+=["### Warnings (review)"]+["- [warn] "+w for w in warns]+[""]
    if notes:  report+=["### Notes"]+["- "+n for n in notes]+[""]
    report.append(EXCEL_SIGNOFF)
    md="\n".join(report)
    rp=os.path.splitext(a.out)[0]+"_normalisation_report.md"
    with open(rp,"w",encoding="utf-8") as f: f.write(md)
    print(md); print("\n[report saved: %s]"%rp)

EXCEL_SIGNOFF="\n".join([
 "## Open in Excel and check (the model's headline numbers only compute in Excel)",
 "1. Open the output, full recalc (Ctrl+Alt+F9).",
 "2. Confirm every asset name in Tenancy Inputs!C matches the Asset Register Cash Flow Output!P5:P30 (toggles 1/0). Mismatches drop units silently.",
 "3. Spot-check 3-4 units: passing rent, ERV, lease start/expiry, rent-review date read as the broker intended; dates show as dates, not text.",
 "4. Check headline outputs: Cash Flow Output IRR/MOIC/profit, Income & Cost Schedule row 36 by-asset income, the Next Buyer CF returns.",
 "5. Confirm no new #REF (only the pre-existing BRE Economics!E75 is expected).",
 "",
 "*Wiring the template into Tenancy Inputs, setting deal assumptions, and reading outputs is the mli-underwrite-runner step - this skill stops at a validated, populated template.*",
])

if __name__=="__main__":
    main()
