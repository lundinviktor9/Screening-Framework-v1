#!/usr/bin/env python3
"""Newbury v3 'Marketing Tenancy Schedule' -> normalised RR. No guarantee (retired).
Vacant@Entry = col G; Break Taken = col K; Vacate@Expiry = col U; UO units (G=N, tenant
'under offer') let from start at the agreed headline rent (col Q)."""
import openpyxl, sys, datetime as dt, argparse
from openpyxl.utils import column_index_from_string as cix
def pdate(x):
    if isinstance(x,dt.datetime): return dt.datetime(x.year,x.month,x.day)
    if isinstance(x,dt.date): return dt.datetime(x.year,x.month,x.day)
    return None
def mny(x): return float(x) if isinstance(x,(int,float)) else None
ap=argparse.ArgumentParser(); ap.add_argument('--src',required=True); ap.add_argument('--asset',required=True)
ap.add_argument('--region',required=True); ap.add_argument('--out',required=True); a=ap.parse_args()
ws=openpyxl.load_workbook(a.src,data_only=True)["Marketing Tenancy Schedule"]
M=dict(unit='D',tenant='E',gia='F',vac='G',start='H',review='I',brkdate='J',brktaken='K',
       expiry='L',passing='O',headline='Q',eyld='S',xyld='T',vacexp='U',ervpsf='V',capex='W',
       void='X',rf='Y',rrmech='Z',epc='AC')
def g(r,k): return ws.cell(r,cix(M[k])).value
hdrmap={'A':'#','B':'Asset Name','C':'Region','D':'Sector','E':'Unit Number','F':'Tenant Name',
  'G':'Area GIA (sq ft)','H':'Entry Yield (NIY)','I':'Exit Yield','J':'Lease Start','K':'Lease Expiry',
  'L':'Break Date','M':'Break Taken (1=Yes,0=No)','N':'Rent Review / MTM Date',
  'O':'Event @ Expiry (Y=Renew / X=Vacate)','P':'Vacant @ Entry (Y/N)','S':'Passing Rent (£ pa)',
  'T':'ERV (£ pa)','U':'ERV (£ psf)','Y':'Assumed Void (mths)','Z':'Assumed Rent Free (mths)',
  'AA':'Re-letting Capex (£ psf)','AY':'1954 Act (Y/N)','AZ':'EPC Rating','BB':'Rent Review (Y/N)','BC':'Term Certain (mths)','BF':'Headline/Guarantee Rent (£pa)'}
out=openpyxl.Workbook(); o=out.active; o.title='RR'
for c,l in hdrmap.items(): o.cell(1,cix(c)).value=l
n=0; flags=[]
for r in range(12, ws.max_row+1):
    u=g(r,'unit')
    if not u or 'Unit' not in str(u): continue
    n+=1; orow=1+n
    tenant=str(g(r,'tenant') or ''); low=tenant.lower()
    vacant=str(g(r,'vac') or '').strip().upper()=='Y'
    uo=('under offer' in low) and not vacant
    brk=g(r,'brktaken') in (1,'1'); vacexp=g(r,'vacexp') in (1,'1')
    relet=(brk or vacexp) and not vacant
    event='X' if (vacant or brk or vacexp) else 'Y'
    area=mny(g(r,'gia')); ervpsf=mny(g(r,'ervpsf'))
    Nrev=pdate(g(r,'review'))
    passing=(mny(g(r,'headline')) if uo else mny(g(r,'passing')))   # UO let from start at agreed rent
    vals={'A':n,'B':a.asset,'C':a.region,'D':'Industrial','E':u,'F':tenant,'G':area,
      'H':mny(g(r,'eyld')),'I':mny(g(r,'xyld')),
      'J':None if vacant else pdate(g(r,'start')),'K':None if vacant else pdate(g(r,'expiry')),
      'L':pdate(g(r,'brkdate')) if brk else None,'M':1 if brk else 0,
      'N':None if (vacant or brk) else Nrev,   # no review for break units
      'O':event,'P':'Y' if vacant else 'N',
      'S':0 if vacant else passing,'T':(ervpsf*area if ervpsf and area else None),'U':ervpsf,
      'Y':(mny(g(r,'void')) if relet else None),'Z':(mny(g(r,'rf')) if relet else None),
      'AA':(mny(g(r,'capex')) if relet else None),
      'AZ':g(r,'epc'),'BB':'Y' if (Nrev and not vacant and not brk) else 'N','BC':60,
      'BF':mny(g(r,'headline'))}  # headline rent -> guarantee/PP pricing (TI U)
    for c,v in vals.items(): o.cell(orow,cix(c)).value=v
    if vacant: flags.append(f"U{n} {u}: VACANT@entry -> ERV cap, deduct 1yr ERV + 20% fee")
    if brk: flags.append(f"U{n} {u}: BREAK taken -> vacate+re-let; NO rent review")
    if vacexp: flags.append(f"U{n} {u}: VACATE@expiry -> re-let")
    if uo: flags.append(f"U{n} {u}: UNDER OFFER -> let from start @ £{passing:,.0f} (agreed headline)")
out.save(a.out)
print(f"Normalised {n} units -> {a.out}\nFLAGS:")
for f in flags: print("  -",f)
