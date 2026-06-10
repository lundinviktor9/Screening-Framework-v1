"""
underwrite/adapter.py - framework-facing wrapper around the vendored MLI engine.

Entry points (see INTERFACE.md):
    run_mode_a(deal_id, normalised_rr_xlsx, rr_sheet, ...) -> Mode A (parse + validate + flag)
    run_mode_b(deal_id, normalised_rr_xlsx, rr_sheet, asset, region, entry, assumptions=None)
        -> Mode B (populate model + headless recalc + verify + returns)

Mode B pipeline: copy the normalised RR sheet into a copy of the PINNED base ->
inject_deal_v21 (header-driven) -> apply assumption dials -> headless LibreOffice recalc ->
verify -> extract returns.

DEPENDENCIES: openpyxl; LibreOffice (`soffice`) on PATH for Mode B recalc. Mode A needs neither.

NOTE (Excel-safety): headless LibreOffice MASKS some Excel errors (coerces text to 0 inside
array formulas). `checks.workbook_error_cells` uses the v21 verify logic; a final Excel
Ctrl+Alt+F9 review remains the gold standard before any number reaches an IC.
"""
from __future__ import annotations
import json
import os
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional


def _soffice_bin() -> str:
    """LibreOffice executable. Resolution order:
      1. SOFFICE_BIN env var (absolute path), if set.
      2. Auto-detect the standard Windows/macOS/Linux install locations — so a normal
         LibreOffice install Just Works with no PATH or env editing.
      3. Fall back to `soffice` on PATH.
    Step 2 exists because Python's subprocess on Windows does not reliably resolve a bare
    `soffice` name from a shell-exported PATH; pointing at the absolute .exe avoids WinError 2."""
    env = os.environ.get("SOFFICE_BIN")
    if env:
        return env
    candidates = (
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/usr/bin/soffice",
        "/usr/local/bin/soffice",
        "/opt/libreoffice/program/soffice",
    )
    for cand in candidates:
        if os.path.exists(cand):
            return cand
    return "soffice"

import openpyxl
from openpyxl.utils import column_index_from_string as cix

ENGINE_DIR = Path(__file__).parent / "engine"
PINNED_BASE = ENGINE_DIR / "base" / "MLI_v21_BASE.xlsx"
INJECTOR = ENGINE_DIR / "inject_deal_v21.py"
VERIFIER = ENGINE_DIR / "verify.py"
ANCHOR_TIEOUT = 8203713.29  # West Craig GA D16 with register Q5=1/Q6=0

# CFO cells -> returns.json keys (deal headline, register Q5=0/Q6=1)
_RETURN_CELLS = {
    "unlevered_irr": ("Cash Flow Output", "L17"),
    "net_investor_irr": ("Cash Flow Output", "K18"),
    "levered_irr": ("Cash Flow Output", "J18"),
    "equity_multiple": ("Cash Flow Output", "J19"),
    "cash_on_cash": ("Cash Flow Output", "L20"),
    "net_exit_price": ("Cash Flow Output", "E5"),
}


class UnderwriteError(RuntimeError):
    pass


# ---------------------------------------------------------------- helpers

def _recalc(xlsx_in: Path, out_dir: Path, profile: Path, timeout: int = 240) -> Path:
    """Force a full LibreOffice recalc-on-load and re-save (faithful round-trip)."""
    (profile / "user").mkdir(parents=True, exist_ok=True)
    reg = profile / "user" / "registrymodifications.xcu"
    if not reg.exists():
        reg.write_text(
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<oor:items xmlns:oor="http://openoffice.org/2001/registry" '
            'xmlns:xs="http://www.w3.org/2001/XMLSchema" '
            'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">\n'
            ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
            '<prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>\n'
            ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
            '<prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>\n'
            '</oor:items>\n'
        )
    out_dir.mkdir(parents=True, exist_ok=True)
    # Path.as_uri() gives a valid file URL on every OS (file:///C:/... on Windows, file:///tmp/...
    # on Linux); the old f"file://{profile}" was malformed on Windows and broke Mode B there.
    profile_uri = profile.resolve().as_uri()
    subprocess.run(
        [_soffice_bin(), f"-env:UserInstallation={profile_uri}", "--headless", "--calc",
         "--convert-to", "xlsx", "--outdir", str(out_dir), str(xlsx_in)],
        check=True, capture_output=True, timeout=timeout,
    )
    out = out_dir / (xlsx_in.stem + ".xlsx")
    if not out.exists():
        raise UnderwriteError(f"recalc produced no output for {xlsx_in}")
    return out


def _copy_rr_into_base(base: Path, rr_xlsx: Path, rr_sheet: str, dest: Path) -> None:
    """Copy the normalised RR sheet from rr_xlsx into a copy of the pinned base, saved to dest."""
    wb = openpyxl.load_workbook(base)
    src = openpyxl.load_workbook(rr_xlsx, data_only=True)
    if rr_sheet not in src.sheetnames:
        raise UnderwriteError(f"RR sheet {rr_sheet!r} not in {rr_xlsx}")
    s = src[rr_sheet]
    if rr_sheet in wb.sheetnames:
        del wb[rr_sheet]
    tgt = wb.create_sheet(rr_sheet)
    for row in s.iter_rows():
        for c in row:
            if c.value is not None:
                tgt.cell(row=c.row, column=c.column, value=c.value)
    wb.save(dest)


def _read_returns(recalced: Path) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(recalced, data_only=True)
    out: Dict[str, Any] = {}
    for key, (sheet, cell) in _RETURN_CELLS.items():
        out[key] = wb[sheet][cell].value
    out["net_purchase_price"] = wb["Global Assumptions"]["D16"].value
    return out


def _count_workbook_errors(recalced: Path) -> int:
    wb = openpyxl.load_workbook(recalced, data_only=True)
    errs = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#NUM!")
    n = 0
    for ws in wb.worksheets:
        if ws.title == "TI Build Guide":
            continue
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value in errs:
                    n += 1
    return n


def _anchor_tieout_ok(prepared_model: Path, profile: Path, workdir: Path) -> bool:
    """Flip the register to West Craig (Q5=1/Q6=0), recalc, check GA D16 == anchor."""
    wb = openpyxl.load_workbook(prepared_model)
    cfo = wb["Cash Flow Output"]; cfo["Q5"] = 1; cfo["Q6"] = 0
    wc = workdir / "anchor_check.xlsx"
    wb.save(wc)
    rec = _recalc(wc, workdir / "anchor_out", profile)
    d16 = openpyxl.load_workbook(rec, data_only=True)["Global Assumptions"]["D16"].value
    return isinstance(d16, (int, float)) and abs(d16 - ANCHOR_TIEOUT) < 0.01


# ---------------------------------------------------------------- assumptions

def _apply_assumptions(wb, a: Optional[Dict[str, Any]], rr_sheet: str) -> List[str]:
    """Write deal-level dials onto the injected model. Returns notes for any that couldn't apply.
    Cells: hold_years->CFO Q26; exit_yield_shift->Q27; rental_growth->GA D49; ltv->CFO Q37;
    scenario (1-4)->CFO Q35. An absolute exit_yield maps to a Q27 shift only when the schedule's
    exit yields are uniform (else it cannot map to a single shift and is flagged)."""
    notes: List[str] = []
    if not a:
        return notes
    cfo = wb["Cash Flow Output"]; ga = wb["Global Assumptions"]
    if a.get("hold_years") is not None:
        cfo["Q26"] = a["hold_years"]                      # D8/D9 (exit) cascade off Q26
    if a.get("exit_yield_shift") is not None:
        cfo["Q27"] = a["exit_yield_shift"]
    elif a.get("exit_yield") is not None:
        ys = set()
        rr = wb[rr_sheet]
        for r in range(2, rr.max_row + 1):
            v = rr.cell(r, cix("I")).value
            if isinstance(v, (int, float)):
                ys.add(round(v, 6))
        if len(ys) == 1:
            cfo["Q27"] = round(a["exit_yield"] - ys.pop(), 6)
        else:
            notes.append(f"exit_yield not applied: schedule exit yields are non-uniform "
                         f"({len(ys)} distinct); set exit_yield_shift instead")
    if a.get("rental_growth") is not None:
        ga["D49"] = a["rental_growth"]
    if a.get("ltv") is not None:
        cfo["Q37"] = a["ltv"]
    if a.get("scenario") is not None:                    # 1-4 (Base UK SPV, ...)
        cfo["Q35"] = a["scenario"]
    for k in ("debt_rate", "tax_scenario", "exit_basis"):
        if a.get(k) is not None:
            notes.append(f"{k} not yet wired into Mode B (documented TODO)")
    return notes


def _apply_entry_yield(prepared: Path, rr_sheet: str, a: Optional[Dict[str, Any]]) -> List[str]:
    """Apply the entry-yield pricing dial BEFORE injection.

    The injector references `RR!H{row}` per unit for Entry Yield (NIY), so an override has to be
    written into column H of the RR sheet in the *prepared* base, ahead of inject_deal_v21.
    (It cannot live in _apply_assumptions, which runs on the model AFTER those refs resolve.)
    A single `entry_yield` is applied uniformly to every unit; leave it null to keep the
    per-unit schedule yields."""
    notes: List[str] = []
    if not a or a.get("entry_yield") is None:
        return notes
    ey = float(a["entry_yield"])
    wb = openpyxl.load_workbook(prepared)
    if rr_sheet not in wb.sheetnames:
        return [f"entry_yield not applied: RR sheet {rr_sheet!r} missing from prepared base"]
    rr = wb[rr_sheet]
    n = 0
    r = 2
    while rr.cell(r, cix("E")).value not in (None, ""):   # E = Unit Number; stop at first blank
        rr.cell(r, cix("H")).value = ey
        n += 1
        r += 1
    wb.save(prepared)
    notes.append(f"entry_yield={ey} applied uniformly to {n} unit(s) (RR col H, pre-injection)")
    return notes


# ---------------------------------------------------------------- Mode B

def run_mode_b(
    deal_id: str,
    normalised_rr_xlsx: str,
    rr_sheet: str,
    asset: str,
    region: str,
    entry: str,                       # YYYY-MM-DD
    assumptions: Optional[Dict[str, Any]] = None,
    base: Optional[str] = None,
    workdir: Optional[str] = None,
) -> Dict[str, Any]:
    """Populate the model from a normalised RR and return the underwrite block."""
    # Diagnostic: print all paths and verify existence
    print(f"\n=== Mode B Diagnostics ===")
    print(f"deal_id={deal_id}, rr_sheet={rr_sheet}")
    print(f"PINNED_BASE={PINNED_BASE}, exists={PINNED_BASE.exists()}")
    print(f"INJECTOR={INJECTOR}, exists={INJECTOR.exists()}")
    print(f"normalised_rr_xlsx={normalised_rr_xlsx}, exists={Path(normalised_rr_xlsx).exists()}")
    print(f"workdir={workdir}")
    print(f"soffice={_soffice_bin()}, exists={os.path.exists(_soffice_bin())}")
    print(f"sys.executable={sys.executable}, exists={os.path.exists(sys.executable)}")
    print(f"=== end diagnostics ===\n", flush=True)

    base_p = Path(base) if base else PINNED_BASE
    if not base_p.exists():
        raise UnderwriteError(f"pinned base not found: {base_p}")
    work = Path(workdir) if workdir else Path(tempfile.mkdtemp(prefix=f"uw_{deal_id}_"))
    work.mkdir(parents=True, exist_ok=True)
    profile = work / "lo_profile"

    prepared = work / "prepared_base.xlsx"
    try:
        _copy_rr_into_base(base_p, Path(normalised_rr_xlsx), rr_sheet, prepared)
    except Exception as e:
        raise UnderwriteError(f"_copy_rr_into_base failed: {e}")

    # Entry-yield pricing dial must be stamped onto RR col H before injection (see helper).
    try:
        entry_yield_notes = _apply_entry_yield(prepared, rr_sheet, assumptions)
    except Exception as e:
        raise UnderwriteError(f"_apply_entry_yield failed: {e}")

    injected = work / "model_injected.xlsx"
    try:
        subprocess.run(
            [sys.executable, str(INJECTOR), "--base", str(prepared), "--rr-sheet", rr_sheet,
             "--asset", asset, "--region", region, "--entry", entry, "--out", str(injected)],
            check=True, capture_output=True, timeout=180,
        )
    except subprocess.CalledProcessError as e:
        raise UnderwriteError(f"injector failed: returncode={e.returncode}, "
                            f"stdout={e.stdout.decode()[:200]}, stderr={e.stderr.decode()[:200]}")
    except Exception as e:
        raise UnderwriteError(f"injector subprocess error: {type(e).__name__}: {e}")

    assumption_notes: List[str] = list(entry_yield_notes)
    if assumptions:
        try:
            wb = openpyxl.load_workbook(injected)
            assumption_notes += _apply_assumptions(wb, assumptions, rr_sheet)
            wb.save(injected)
        except Exception as e:
            raise UnderwriteError(f"_apply_assumptions failed: {e}")

    try:
        recalced = _recalc(injected, work / "recalc_out", profile)
    except Exception as e:
        raise UnderwriteError(f"LibreOffice recalc failed: {type(e).__name__}: {e}")

    try:
        returns = _read_returns(recalced)
        err_cells = _count_workbook_errors(recalced)
        tieout_ok = _anchor_tieout_ok(injected, profile, work)
    except Exception as e:
        raise UnderwriteError(f"post-recalc processing failed: {e}")

    checks = {
        "pass": bool(tieout_ok and err_cells == 0),
        "anchor_tieout_ok": tieout_ok,
        "workbook_error_cells": err_cells,
    }
    return {
        "deal_id": deal_id,
        "mode": "B",
        "model_xlsx": str(recalced),
        "returns": returns,
        "checks": checks,
        "assumption_notes": assumption_notes,
    }


# ---------------------------------------------------------------- Mode A

def classify_flags(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Deterministic signal -> flag. The LLM 'classify' refinement is a future hook."""
    flags: List[Dict[str, Any]] = []
    for r in rows:
        unit = r.get("Unit Number") or r.get("E") or "?"
        tenant = str(r.get("Tenant Name") or "").lower()
        vacant = str(r.get("Vacant @ Entry (Y/N)") or r.get("P") or "").upper() == "Y"
        if vacant:
            flags.append({"unit": unit, "signal": "vacant@entry",
                          "treatment": "capitalise guarantee/headline rent; re-let at void/RF defaults",
                          "needs_signoff": True})
        if "under offer" in tenant or tenant == "uo":
            flags.append({"unit": unit, "signal": "under_offer",
                          "treatment": "let from agreed start on agreed rent/term",
                          "needs_signoff": True})
        if str(r.get("Break Taken (1=Yes,0=No)") or r.get("M") or "") in ("1", "1.0"):
            flags.append({"unit": unit, "signal": "break_taken",
                          "treatment": "break exercised -> vacate + re-let; suppress post-break review",
                          "needs_signoff": True})
    return flags


def run_mode_a(
    deal_id: str,
    normalised_rr_xlsx: str,
    rr_sheet: str,
    schema_dir: Optional[str] = None,
) -> Dict[str, Any]:
    """Parse + validate the normalised RR against the schema and surface judgement-call flags."""
    import re as _re
    sdir = Path(schema_dir) if schema_dir else Path(__file__).parent / "schemas"
    schema = json.loads((sdir / "tenancy_schedule.schema.json").read_text(encoding="utf-8"))
    required = [c["name"] for c in schema["columns"] if c.get("required")]

    wb = openpyxl.load_workbook(normalised_rr_xlsx, data_only=True)
    s = wb[rr_sheet]
    headers = {(s.cell(1, c).value or ""): c for c in range(1, s.max_column + 1)}
    rows: List[Dict[str, Any]] = []
    for r in range(2, s.max_row + 1):
        if not s.cell(r, cix("E")).value:  # Unit Number empty -> end of data
            continue
        rows.append({h: s.cell(r, col).value for h, col in headers.items() if h})

    # RR headers contain currency symbols / varied spacing; schema names avoid them. Compare on a
    # normalised key (alphanumerics only) so "Passing Rent (pa)" matches "Passing Rent (GBP pa)".
    def _norm(x: str) -> str:
        return _re.sub(r"[^a-z0-9]", "", str(x).lower())
    have = {_norm(h) for h in headers}
    schema_errors = [f"required field has no column: {f}"
                     for f in required if _norm(f) not in have]

    return {
        "deal_id": deal_id,
        "mode": "A",
        "units": len(rows),
        "flags": classify_flags(rows),
        "schema_errors": schema_errors,
    }


if __name__ == "__main__":  # tiny CLI for manual runs
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("mode", choices=["A", "B"])
    ap.add_argument("--deal-id", required=True)
    ap.add_argument("--rr-xlsx", required=True)
    ap.add_argument("--rr-sheet", required=True)
    ap.add_argument("--asset"); ap.add_argument("--region"); ap.add_argument("--entry")
    a = ap.parse_args()
    if a.mode == "A":
        print(json.dumps(run_mode_a(a.deal_id, a.rr_xlsx, a.rr_sheet), indent=2, default=str))
    else:
        print(json.dumps(run_mode_b(a.deal_id, a.rr_xlsx, a.rr_sheet, a.asset, a.region, a.entry),
                         indent=2, default=str))
